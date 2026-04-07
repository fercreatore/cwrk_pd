#!/usr/bin/env python3
"""
Insertar pedido OLK/Olympikus - Abr-May 2026 (GLOBAL BRANDS SA, proveedor 722)
Empresa: H4 -> INSERT en MSGESTION03.dbo.pedico2 + pedico1

Fuente: /Volumes/compartido/COMPRAS/Pedidos Invierno/PEDIDO OLK ABRIL-MAYO.xlsx
Sheet "NP Q1Q2", headers row 14, data rows 15-137
Solo filas con columna AB (Mod) > 0

8 items confirmados, 156 pares, ~$6.324.240

EJECUTAR:
  python3 _scripts_oneshot/insertar_olk_olympikus.py --dry-run
  python3 _scripts_oneshot/insertar_olk_olympikus.py --ejecutar
"""

import sys
import os
import pyodbc
import openpyxl
from datetime import date

# ── CONFIG ────────────────────────────────────────────────
# Fix SSL para Mac (OpenSSL 3.x + SQL Server 2012 TLS 1.0)
import platform
if platform.system() != "Windows":
    _ssl_conf = "/tmp/openssl_legacy.cnf"
    if not os.path.exists(_ssl_conf):
        with open(_ssl_conf, "w") as _f:
            _f.write(
                "openssl_conf = openssl_init\n"
                "[openssl_init]\nssl_conf = ssl_sect\n"
                "[ssl_sect]\nsystem_default = system_default_sect\n"
                "[system_default_sect]\n"
                "MinProtocol = TLSv1\nCipherString = DEFAULT@SECLEVEL=0\n"
            )
    os.environ.setdefault("OPENSSL_CONF", _ssl_conf)

# INSERT en MSGESTION03 (H4)
CONN_STR = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=192.168.2.111;DATABASE=msgestion03;"
    "UID=am;PWD=dl;TrustServerCertificate=yes;Encrypt=no"
)

# Articulos: lookup en msgestion01art
CONN_ART = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=192.168.2.111;DATABASE=msgestion01art;"
    "UID=am;PWD=dl;TrustServerCertificate=yes;Encrypt=no"
)

PROVEEDOR = 722
DENOMINACION = "GLOBAL BRANDS S.A."
EMPRESA = "H4"
OBS = "Pedido OLK Olympikus Abr-May 2026. 156 pares. $6.324.240. GLOBAL BRANDS SA"
FECHA = date.today().strftime('%Y%m%d')

# Excel path
EXCEL_PATH = "/Volumes/compartido/COMPRAS/Pedidos Invierno/PEDIDO OLK ABRIL-MAYO.xlsx"
SHEET_NAME = "NP Q1Q2"
HEADER_ROW = 14
DATA_START = 15
DATA_END = 137  # inclusive

# Talle columns: K=11 (talle 34) through Y=25 (talle 48)
# Column indices (1-based): K=11, L=12, ..., Y=25
TALLE_COL_START = 11  # col K = talle 34
TALLE_COL_END = 25    # col Y = talle 48
TALLES = list(range(34, 49))  # 34, 35, ..., 48

# Key columns (1-based)
COL_CODEQUIS = 2   # B: CodEquis (11-digit OLK code)
COL_MODELO = 5     # E: MODELO
COL_PCOM = 6       # F: Pcom (purchase price)
COL_COLOR = 9      # I: Color
COL_MOD = 28       # AB: Mod (modules, filter > 0)
COL_PARES = 29     # AC: Pares total

EXPECTED_PARES = 156
EXPECTED_MONTO = 6_324_240


def leer_excel(path):
    """Lee el Excel y retorna las filas con Mod > 0."""
    print(f"Leyendo Excel: {path}")
    print(f"  Sheet: {SHEET_NAME}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    items = []
    for row_num in range(DATA_START, DATA_END + 1):
        mod_val = ws.cell(row=row_num, column=COL_MOD).value
        if mod_val is None or (isinstance(mod_val, (int, float)) and mod_val <= 0):
            continue

        codequis = str(ws.cell(row=row_num, column=COL_CODEQUIS).value or "").strip()
        modelo = str(ws.cell(row=row_num, column=COL_MODELO).value or "").strip()
        pcom = ws.cell(row=row_num, column=COL_PCOM).value or 0
        color = str(ws.cell(row=row_num, column=COL_COLOR).value or "").strip()
        pares_total = ws.cell(row=row_num, column=COL_PARES).value or 0

        # Read talle quantities
        talle_qty = {}
        for i, talle in enumerate(TALLES):
            col_idx = TALLE_COL_START + i
            qty = ws.cell(row=row_num, column=col_idx).value
            if qty and isinstance(qty, (int, float)) and qty > 0:
                talle_qty[talle] = int(qty)

        items.append({
            "row": row_num,
            "codequis": codequis,
            "modelo": modelo,
            "pcom": float(pcom),
            "color": color,
            "mod": float(mod_val),
            "pares_total": int(pares_total) if pares_total else sum(talle_qty.values()),
            "talle_qty": talle_qty,
        })

    wb.close()
    print(f"  Filas con Mod > 0: {len(items)}")
    return items


def buscar_articulos_por_sinonimo(items):
    """
    Para cada item, busca en msgestion01art.dbo.articulo los articulos
    que matcheen por sinonimo LIKE codequis%.
    El sinonimo en la DB tiene formato: codequis + 2 digitos de talle.
    """
    conn = pyodbc.connect(CONN_ART, timeout=15)
    cur = conn.cursor()

    resultados = []
    errores = []

    for item in items:
        codequis = item["codequis"]
        if not codequis:
            errores.append(f"  Row {item['row']}: CodEquis vacio para {item['modelo']} {item['color']}")
            continue

        # Buscar todos los articulos cuyo sinonimo empiece con el CodEquis
        cur.execute("""
            SELECT codigo, codigo_sinonimo, descripcion_1
            FROM articulo
            WHERE codigo_sinonimo LIKE ? + '%'
              AND estado = 'V'
            ORDER BY codigo_sinonimo
        """, codequis)

        rows = cur.fetchall()
        if not rows:
            errores.append(f"  Row {item['row']}: Sin match para CodEquis={codequis} ({item['modelo']} {item['color']})")
            continue

        # Build sinonimo -> (codigo, denominacion) map
        sino_map = {}
        for r in rows:
            sino = (r.codigo_sinonimo or "").strip()
            sino_map[sino] = (r.codigo, (r.descripcion_1 or "").strip())

        # For each talle with qty > 0, find the matching article
        for talle, qty in sorted(item["talle_qty"].items()):
            # Try common sinonimo patterns:
            #   codequis + "XX" (2-digit talle)
            #   codequis + "0XX" (3-digit with leading zero)
            talle_str = str(talle)
            talle_str_2 = f"{talle:02d}"

            # Pattern 1: codequis + talle (e.g., "72256590001" + "38")
            sino_try = codequis + talle_str
            if sino_try in sino_map:
                cod_art, denom = sino_map[sino_try]
                resultados.append({
                    "articulo": cod_art,
                    "sinonimo": sino_try,
                    "descripcion": f"{denom} T{talle}",
                    "cantidad": qty,
                    "precio": item["pcom"],
                    "modelo": item["modelo"],
                    "color": item["color"],
                    "talle": talle,
                })
                continue

            # Pattern 2: codequis + "0" + talle
            sino_try2 = codequis + "0" + talle_str
            if sino_try2 in sino_map:
                cod_art, denom = sino_map[sino_try2]
                resultados.append({
                    "articulo": cod_art,
                    "sinonimo": sino_try2,
                    "descripcion": f"{denom} T{talle}",
                    "cantidad": qty,
                    "precio": item["pcom"],
                    "modelo": item["modelo"],
                    "color": item["color"],
                    "talle": talle,
                })
                continue

            # Pattern 3: Try matching last 2 digits of sinonimo with talle
            found = False
            for sino_key, (cod_art, denom) in sino_map.items():
                if sino_key.endswith(talle_str_2) and len(sino_key) >= len(codequis) + 2:
                    resultados.append({
                        "articulo": cod_art,
                        "sinonimo": sino_key,
                        "descripcion": f"{denom} T{talle}",
                        "cantidad": qty,
                        "precio": item["pcom"],
                        "modelo": item["modelo"],
                        "color": item["color"],
                        "talle": talle,
                    })
                    found = True
                    break

            if not found:
                errores.append(
                    f"  Row {item['row']}: Talle {talle} sin match en sinonimos para "
                    f"CodEquis={codequis} ({item['modelo']} {item['color']}). "
                    f"Sinonimos encontrados: {list(sino_map.keys())[:5]}"
                )

    conn.close()
    return resultados, errores


def main():
    dry_run = True
    if len(sys.argv) > 1:
        if sys.argv[1] == "--ejecutar":
            dry_run = False
        elif sys.argv[1] != "--dry-run":
            print("Uso: python insertar_olk_olympikus.py [--dry-run | --ejecutar]")
            sys.exit(1)

    print("=" * 70)
    print(f"PEDIDO OLK/OLYMPIKUS - GLOBAL BRANDS SA (prov {PROVEEDOR})")
    print(f"Empresa: {EMPRESA} -> MSGESTION03.dbo.pedico2/pedico1")
    print(f"Modo: {'DRY RUN' if dry_run else 'EJECUCION REAL'}")
    print("=" * 70)

    # 1. Leer Excel
    if not os.path.exists(EXCEL_PATH):
        print(f"\nERROR: No se encuentra el Excel:\n  {EXCEL_PATH}")
        print("Montar el compartido primero o copiar el archivo.")
        sys.exit(1)

    items = leer_excel(EXCEL_PATH)
    if not items:
        print("ERROR: No se encontraron items con Mod > 0")
        sys.exit(1)

    # Show items found
    print(f"\n{'─' * 70}")
    print("ITEMS DEL EXCEL:")
    for item in items:
        pares_check = sum(item["talle_qty"].values())
        print(f"  {item['modelo']:30s} {item['color']:20s} "
              f"Pcom=${item['pcom']:>12,.0f}  Pares={pares_check:>3d}  "
              f"CodEquis={item['codequis']}")
        talles_str = ", ".join(f"T{t}x{q}" for t, q in sorted(item["talle_qty"].items()))
        print(f"    Talles: {talles_str}")

    total_pares_excel = sum(sum(it["talle_qty"].values()) for it in items)
    total_monto_excel = sum(
        sum(it["talle_qty"].values()) * it["pcom"] for it in items
    )
    print(f"\n  Total pares (Excel): {total_pares_excel}")
    print(f"  Total monto (Excel): ${total_monto_excel:,.0f}")

    # 2. Match articles by sinonimo
    print(f"\n{'─' * 70}")
    print("BUSCANDO ARTICULOS POR SINONIMO...")
    detalle, errores = buscar_articulos_por_sinonimo(items)

    if errores:
        print(f"\n*** ERRORES DE MATCH ({len(errores)}) ***")
        for err in errores:
            print(err)

    if not detalle:
        print("\nERROR: No se encontro ningun articulo. Verificar sinonimos en la DB.")
        sys.exit(1)

    # 3. Summary
    total_pares = sum(d["cantidad"] for d in detalle)
    total_monto = sum(d["cantidad"] * d["precio"] for d in detalle)

    print(f"\n{'─' * 70}")
    print(f"DETALLE A INSERTAR: {len(detalle)} renglones")
    print(f"  Pares: {total_pares} (esperado: {EXPECTED_PARES})")
    print(f"  Monto: ${total_monto:,.0f} (esperado: ${EXPECTED_MONTO:,.0f})")
    print()

    for d in detalle:
        print(f"  [{d['articulo']:>6d}] {d['sinonimo']:20s} {d['descripcion'][:40]:40s} "
              f"x{d['cantidad']:>2d}  ${d['precio']:>12,.0f}")

    if total_pares != EXPECTED_PARES:
        print(f"\n*** ADVERTENCIA: Pares={total_pares} != esperado {EXPECTED_PARES} ***")
        print("    Revisar errores de match arriba.")

    # 4. DRY RUN / EJECUTAR
    if dry_run:
        print(f"\n{'─' * 70}")
        print("[DRY RUN] Ningún dato fue escrito. Usar --ejecutar para insertar.")
        return

    # Confirmation
    print(f"\n{'─' * 70}")
    resp = input(f"Confirmar INSERT de {total_pares} pares en MSGESTION03? (S/N): ")
    if resp.strip().upper() != 'S':
        print("Cancelado.")
        return

    # 5. INSERT
    conn = pyodbc.connect(CONN_STR, timeout=15)
    conn.autocommit = False
    cur = conn.cursor()

    try:
        # Get next numero and orden
        cur.execute(
            "SELECT ISNULL(MAX(numero), 0) + 1 FROM pedico2 WHERE codigo = 8"
        )
        numero = cur.fetchone()[0]

        cur.execute(
            "SELECT ISNULL(MAX(orden), 0) + 1 FROM pedico2 WHERE codigo = 8"
        )
        orden = cur.fetchone()[0]
        if orden > 99:
            orden = 1

        print(f"\nInsertando pedido: numero={numero}, orden={orden}")

        # Verify it doesn't exist
        cur.execute("""
            SELECT COUNT(*) FROM pedico2
            WHERE codigo=8 AND letra='X' AND sucursal=1 AND numero=?
        """, numero)
        if cur.fetchone()[0] > 0:
            print(f"ERROR: Pedido {numero} ya existe!")
            conn.close()
            return

        # INSERT pedico2 (cabecera)
        cur.execute("""
            INSERT INTO pedico2 (
                codigo, letra, sucursal, numero, orden,
                deposito, cuenta, denominacion,
                fecha_comprobante, fecha_proceso,
                descuento_general, monto_descuento,
                bonificacion_general, monto_bonificacion,
                financiacion_general, monto_financiacion,
                iva1, monto_iva1, iva2, monto_iva2,
                monto_impuesto, monto_exento,
                importe_neto,
                estado, condicion_iva,
                copias, usuario,
                campo, sistema_cc, moneda, sector,
                forma_pago, tipo_vcto_pago, tipo_operacion, tipo_ajuste,
                medio_pago, cuenta_cc,
                plan_canje, cuenta_y_orden, pack, reintegro, cambio, transferencia,
                concurso, entregador,
                observaciones
            ) VALUES (
                8, 'X', 1, ?, ?,
                0, ?, ?,
                CONVERT(datetime, ?, 112), GETDATE(),
                0, 0,
                0, 0,
                0, 0,
                21, 0, 10.5, 0,
                0, 0,
                0,
                'V', 'I',
                1, 'COWORK',
                0, 2, 0, 0,
                0, 0, 0, 0,
                ' ', ?,
                'N', 'N', 'N', 'N', 'N', 'N',
                'N', 0,
                ?
            )
        """, numero, orden,
            PROVEEDOR, DENOMINACION,
            FECHA,
            PROVEEDOR,
            OBS)

        # INSERT pedico1 (detalle)
        for renglon, d in enumerate(detalle, 1):
            cur.execute("""
                INSERT INTO pedico1 (
                    codigo, letra, sucursal, numero, orden, renglon,
                    articulo, descripcion, precio, cantidad,
                    descuento_reng1, descuento_reng2,
                    estado, fecha,
                    cuenta,
                    codigo_sinonimo
                ) VALUES (
                    8, 'X', 1, ?, ?, ?,
                    ?, ?, ?, ?,
                    0, 0,
                    'V', CONVERT(datetime, ?, 112),
                    ?,
                    ?
                )
            """, numero, orden, renglon,
                d["articulo"], d["descripcion"], d["precio"], d["cantidad"],
                FECHA,
                PROVEEDOR,
                d["sinonimo"])

        conn.commit()
        print(f"\nPEDIDO INSERTADO EXITOSAMENTE:")
        print(f"  Numero:     {numero}")
        print(f"  Orden:      {orden}")
        print(f"  Proveedor:  {PROVEEDOR} - {DENOMINACION}")
        print(f"  Pares:      {total_pares}")
        print(f"  Monto:      ${total_monto:,.0f}")
        print(f"  Renglones:  {len(detalle)}")

        # Post-insert verification
        cur.execute(
            "SELECT COUNT(*), SUM(cantidad) FROM pedico1 "
            "WHERE codigo=8 AND numero=? AND orden=?",
            numero, orden
        )
        vrow = cur.fetchone()
        print(f"\nVerificacion DB: {vrow[0]} renglones, {vrow[1]} pares en pedico1")

    except Exception as e:
        conn.rollback()
        print(f"\nERROR - ROLLBACK: {e}")
        raise
    finally:
        conn.close()


if __name__ == '__main__':
    main()
