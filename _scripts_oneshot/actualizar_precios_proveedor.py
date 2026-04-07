#!/usr/bin/env python3
"""
actualizar_precios_proveedor.py — Actualiza precio_fabrica + recalcula cadena
=============================================================================
Para cuando llega una nueva lista de precios de un proveedor.

Dos modos:
  1. Por porcentaje: "Aumentar todos los precios de TOPPER un 15%"
  2. Por Excel: "Actualizar precios según la lista nueva del proveedor"

Uso:
  # Aumento porcentual para un proveedor
  python actualizar_precios_proveedor.py --proveedor 668 --aumento 15

  # Aumento porcentual para una marca
  python actualizar_precios_proveedor.py --marca 314 --aumento 15

  # Desde Excel (col A=codigo o sinonimo, col B=precio_nuevo)
  python actualizar_precios_proveedor.py --excel lista_precios.xlsx

  # Opciones adicionales
  --ejecutar          # Aplicar cambios (default: dry run)
  --csv               # Exportar análisis
  --solo-vigentes     # Solo estado='V' (default)
  --incluir-todos     # Incluir estado='U' también

IMPORTANTE: Actualiza msgestion01art.dbo.articulo (tabla compartida H4+CLZ)
"""

import sys
import os
import socket
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pyodbc

# ── Conexión ──
_hostname = socket.gethostname().upper()
if _hostname in ("DELL-SVR", "DELLSVR"):
    SERVIDOR = "localhost"
else:
    SERVIDOR = "192.168.2.111"

CONN_STR = (
    f"DRIVER={{ODBC Driver 17 for SQL Server}};"
    f"SERVER={SERVIDOR};"
    f"DATABASE=msgestion01art;"
    f"UID=am;PWD=dl;"
    f"TrustServerCertificate=yes;Encrypt=no;"
)


def recalcular_cadena(precio_fabrica, descuento, utilidad_1, utilidad_2, utilidad_3, utilidad_4):
    """Recalcula precio_costo y precio_1..4 desde precio_fabrica."""
    precio_costo = round(precio_fabrica * (1 - descuento / 100), 2)
    return {
        "precio_costo": precio_costo,
        "precio_sugerido": precio_costo,
        "precio_1": round(precio_costo * (1 + utilidad_1 / 100), 2),
        "precio_2": round(precio_costo * (1 + utilidad_2 / 100), 2),
        "precio_3": round(precio_costo * (1 + utilidad_3 / 100), 2),
        "precio_4": round(precio_costo * (1 + utilidad_4 / 100), 2),
    }


def aumento_porcentual(cursor, conn, filtro_sql, filtro_params, pct_aumento, dry_run=True):
    """
    Aumenta precio_fabrica en un % y recalcula toda la cadena.
    Usa las utilidades que YA están en cada artículo.
    """
    estados = "('V')"  # solo vigentes por default
    if "--incluir-todos" in sys.argv:
        estados = "('V','U')"

    sql = f"""
        SELECT codigo, descripcion_1, proveedor, marca,
               precio_fabrica, precio_costo, precio_1, precio_2, precio_3, precio_4,
               descuento, utilidad_1, utilidad_2, utilidad_3, utilidad_4
        FROM articulo
        WHERE estado IN {estados}
          AND precio_fabrica > 0
          AND marca NOT IN (1316, 1317, 1158, 436)
          AND {filtro_sql}
        ORDER BY codigo
    """
    cursor.execute(sql, filtro_params)
    articulos = cursor.fetchall()

    print(f"\nArtículos encontrados: {len(articulos)}")
    if not articulos:
        print("No se encontraron artículos con ese filtro.")
        return

    cambios = []
    for art in articulos:
        precio_fab_actual = float(art.precio_fabrica)
        precio_fab_nuevo = round(precio_fab_actual * (1 + pct_aumento / 100), 2)

        desc = float(art.descuento or 0)
        u1 = float(art.utilidad_1 or 0)
        u2 = float(art.utilidad_2 or 0)
        u3 = float(art.utilidad_3 or 0)
        u4 = float(art.utilidad_4 or 0)

        # Si no tiene utilidades, mantener precios proporcionales
        if u1 == 0 and u2 == 0:
            # Calcular utilidades implícitas desde los precios actuales
            p_costo_actual = float(art.precio_costo or 0)
            if p_costo_actual > 0:
                p1_actual = float(art.precio_1 or 0)
                u1_impl = ((p1_actual / p_costo_actual) - 1) * 100 if p1_actual > 0 else 100
                u2_impl = ((float(art.precio_2 or 0) / p_costo_actual) - 1) * 100 if float(art.precio_2 or 0) > 0 else 124
                u3_impl = ((float(art.precio_3 or 0) / p_costo_actual) - 1) * 100 if float(art.precio_3 or 0) > 0 else 60
                u4_impl = ((float(art.precio_4 or 0) / p_costo_actual) - 1) * 100 if float(art.precio_4 or 0) > 0 else 45
                u1, u2, u3, u4 = u1_impl, u2_impl, u3_impl, u4_impl
            else:
                u1, u2, u3, u4 = 100, 124, 60, 45

        nuevo = recalcular_cadena(precio_fab_nuevo, desc, u1, u2, u3, u4)

        cambios.append({
            "codigo": art.codigo,
            "desc": (art.descripcion_1 or "").strip()[:45],
            "p_fab_ant": precio_fab_actual,
            "p_fab_new": precio_fab_nuevo,
            "p1_ant": float(art.precio_1 or 0),
            "p1_new": nuevo["precio_1"],
            "nuevo": nuevo,
        })

    # Mostrar resumen
    print(f"\n{'Codigo':>8} {'Descripcion':45} {'P.Fab Ant':>10} {'P.Fab New':>10} {'P1 Ant':>10} {'P1 New':>10}")
    print("-" * 100)
    for c in cambios[:30]:
        print(f"{c['codigo']:>8} {c['desc']:45} "
              f"${c['p_fab_ant']:>9,.0f} ${c['p_fab_new']:>9,.0f} "
              f"${c['p1_ant']:>9,.0f} ${c['p1_new']:>9,.0f}")
    if len(cambios) > 30:
        print(f"  ... y {len(cambios) - 30} más")

    total_arts = len(cambios)
    p1_promedio_ant = sum(c['p1_ant'] for c in cambios) / total_arts
    p1_promedio_new = sum(c['p1_new'] for c in cambios) / total_arts

    print(f"\nResumen:")
    print(f"  Artículos a actualizar: {total_arts}")
    print(f"  Aumento precio fábrica: {pct_aumento:+.1f}%")
    print(f"  Precio 1 promedio: ${p1_promedio_ant:,.0f} → ${p1_promedio_new:,.0f} ({(p1_promedio_new/p1_promedio_ant - 1)*100:+.1f}%)")

    if dry_run:
        print(f"\n[DRY RUN] No se aplicó ningún cambio.")
        print(f"Para ejecutar: agregar --ejecutar")
        return

    resp = input(f"\nConfirmar UPDATE de {total_arts} artículos? (S/N): ").strip().upper()
    if resp != 'S':
        print("Cancelado.")
        return

    update_sql = """
        UPDATE articulo
        SET precio_fabrica = ?,
            precio_costo = ?,
            precio_sugerido = ?,
            precio_1 = ?,
            precio_2 = ?,
            precio_3 = ?,
            precio_4 = ?
        WHERE codigo = ?
    """

    updated = 0
    for c in cambios:
        n = c['nuevo']
        cursor.execute(update_sql, (
            c['p_fab_new'],
            n['precio_costo'], n['precio_sugerido'],
            n['precio_1'], n['precio_2'], n['precio_3'], n['precio_4'],
            c['codigo'],
        ))
        updated += 1

    conn.commit()
    print(f"\n{updated} artículos actualizados exitosamente.")
    print(f"Backup: los precios anteriores se pueden reconstruir dividiendo por {1 + pct_aumento/100:.4f}")


def desde_excel(cursor, conn, excel_path, dry_run=True):
    """
    Actualiza precios desde un Excel con columnas:
      Col A: codigo ERP o codigo_sinonimo
      Col B: precio_fabrica nuevo
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: pip install openpyxl")
        return

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    lineas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        codigo_o_sino = str(row[0]).strip()
        precio_nuevo = float(row[1])
        if precio_nuevo <= 0:
            continue
        lineas.append((codigo_o_sino, precio_nuevo))

    wb.close()
    print(f"\nLíneas leídas del Excel: {len(lineas)}")

    cambios = []
    no_encontrados = []

    for codigo_o_sino, precio_fab_nuevo in lineas:
        # Buscar por codigo o sinonimo
        cursor.execute("""
            SELECT codigo, descripcion_1, proveedor,
                   precio_fabrica, precio_1,
                   descuento, utilidad_1, utilidad_2, utilidad_3, utilidad_4
            FROM articulo
            WHERE (CAST(codigo AS VARCHAR) = ? OR codigo_sinonimo = ?)
              AND estado = 'V'
        """, codigo_o_sino, codigo_o_sino)

        rows = cursor.fetchall()
        if not rows:
            no_encontrados.append(codigo_o_sino)
            continue

        for art in rows:
            desc = float(art.descuento or 0)
            u1 = float(art.utilidad_1 or 100)
            u2 = float(art.utilidad_2 or 124)
            u3 = float(art.utilidad_3 or 60)
            u4 = float(art.utilidad_4 or 45)

            nuevo = recalcular_cadena(precio_fab_nuevo, desc, u1, u2, u3, u4)

            cambios.append({
                "codigo": art.codigo,
                "desc": (art.descripcion_1 or "").strip()[:45],
                "p_fab_ant": float(art.precio_fabrica or 0),
                "p_fab_new": precio_fab_nuevo,
                "p1_ant": float(art.precio_1 or 0),
                "p1_new": nuevo["precio_1"],
                "nuevo": nuevo,
            })

    if no_encontrados:
        print(f"\nNo encontrados ({len(no_encontrados)}):")
        for nf in no_encontrados[:10]:
            print(f"  - {nf}")

    if not cambios:
        print("No hay cambios para aplicar.")
        return

    # Mostrar y aplicar (misma lógica que aumento_porcentual)
    print(f"\n{'Codigo':>8} {'Descripcion':45} {'P.Fab Ant':>10} {'P.Fab New':>10} {'P1 Ant':>10} {'P1 New':>10}")
    print("-" * 100)
    for c in cambios[:30]:
        print(f"{c['codigo']:>8} {c['desc']:45} "
              f"${c['p_fab_ant']:>9,.0f} ${c['p_fab_new']:>9,.0f} "
              f"${c['p1_ant']:>9,.0f} ${c['p1_new']:>9,.0f}")

    print(f"\nTotal: {len(cambios)} artículos a actualizar")

    if dry_run:
        print(f"\n[DRY RUN] No se aplicó ningún cambio.")
        return

    resp = input(f"\nConfirmar? (S/N): ").strip().upper()
    if resp != 'S':
        print("Cancelado.")
        return

    update_sql = """
        UPDATE articulo
        SET precio_fabrica = ?,
            precio_costo = ?, precio_sugerido = ?,
            precio_1 = ?, precio_2 = ?, precio_3 = ?, precio_4 = ?
        WHERE codigo = ?
    """
    for c in cambios:
        n = c['nuevo']
        cursor.execute(update_sql, (
            c['p_fab_new'],
            n['precio_costo'], n['precio_sugerido'],
            n['precio_1'], n['precio_2'], n['precio_3'], n['precio_4'],
            c['codigo'],
        ))

    conn.commit()
    print(f"\n{len(cambios)} artículos actualizados.")


def main():
    dry_run = "--ejecutar" not in sys.argv

    print("=" * 80)
    print("ACTUALIZAR PRECIOS PROVEEDOR")
    print(f"Modo: {'DRY RUN' if dry_run else 'EJECUCIÓN REAL'}")
    print(f"Servidor: {SERVIDOR}")
    print("=" * 80)

    conn = pyodbc.connect(CONN_STR, timeout=15)
    cursor = conn.cursor()

    # Modo Excel
    for i, arg in enumerate(sys.argv):
        if arg == "--excel" and i + 1 < len(sys.argv):
            desde_excel(cursor, conn, sys.argv[i + 1], dry_run)
            conn.close()
            return

    # Modo porcentual
    pct_aumento = None
    filtro_sql = "1=1"
    filtro_params = []

    for i, arg in enumerate(sys.argv):
        if arg == "--aumento" and i + 1 < len(sys.argv):
            pct_aumento = float(sys.argv[i + 1])
        if arg == "--proveedor" and i + 1 < len(sys.argv):
            filtro_sql = "proveedor = ?"
            filtro_params = [int(sys.argv[i + 1])]
        if arg == "--marca" and i + 1 < len(sys.argv):
            filtro_sql = "marca = ?"
            filtro_params = [int(sys.argv[i + 1])]

    if pct_aumento is None:
        print("\nUso:")
        print("  python actualizar_precios_proveedor.py --proveedor 668 --aumento 15")
        print("  python actualizar_precios_proveedor.py --marca 314 --aumento 15")
        print("  python actualizar_precios_proveedor.py --excel lista_precios.xlsx")
        print("\nOpciones: --ejecutar, --csv, --incluir-todos")
        conn.close()
        return

    if filtro_sql == "1=1":
        print("\nERROR: Especificar --proveedor o --marca para evitar actualizar TODO.")
        print("Si realmente querés actualizar todo, usá auto_markup.py")
        conn.close()
        return

    aumento_porcentual(cursor, conn, filtro_sql, filtro_params, pct_aumento, dry_run)
    conn.close()


if __name__ == '__main__':
    main()
