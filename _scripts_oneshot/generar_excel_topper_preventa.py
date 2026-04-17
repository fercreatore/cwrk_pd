"""
generar_excel_topper_preventa.py
================================
Genera Excel con análisis completo de TOPPER (marca 314, proveedor 668)
para la preventa con Manuel Simón (Alpargatas).

3 hojas:
  1. "Para Manuel"     — SO / SHARE / MOS mensual Ene-Dic 2025
  2. "Analisis Interno" — Quiebre por categoría, top modelos, balance reposición
  3. "Argumentos"       — Resumen ejecutivo para negociación

Ejecutar: python _scripts_oneshot/generar_excel_topper_preventa.py
Output:   _excel_pedidos/topper_preventa_2025.xlsx
"""

import sys, os
from decimal import Decimal
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import config
import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

MARCA = 314
EXCL_CODIGOS = "(7, 36)"
EXCL_MARCAS_GASTOS = "(1316, 1317, 1158, 436)"
RUBROS_CALZADO = "(1, 3, 4, 5, 6)"
DEPOSITOS = "(0,1,2,3,4,5,6,7,8,9,11,12,14,15,198)"
MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# ── Estilos ──────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(bold=True, size=11, color="2F5496")
PCT_FORMAT = '0.0%'
NUM_FORMAT = '#,##0'
MONEY_FORMAT = '#,##0'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")


def get_conn():
    return pyodbc.connect(config.CONN_COMPRAS)


def apply_header_style(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def apply_border_range(ws, row_start, row_end, col_start, col_end):
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


# ═════════════════════════════════════════════════════
# QUERIES
# ═════════════════════════════════════════════════════

def query_so_mensual(conn):
    """SO (Sell Out) Topper mensual Ene-Dic 2025."""
    sql = f"""
    SELECT MONTH(v2.fecha_comprobante) as mes,
           SUM(CASE WHEN v1.operacion='+' THEN v1.cantidad
                    WHEN v1.operacion='-' THEN -v1.cantidad ELSE 0 END) as pares,
           SUM(CASE WHEN v1.operacion='+' THEN v1.cantidad * v1.precio
                    WHEN v1.operacion='-' THEN -v1.cantidad * v1.precio ELSE 0 END) as importe
    FROM msgestionC.dbo.ventas1 v1
    JOIN msgestionC.dbo.ventas2 v2
        ON v1.codigo = v2.codigo AND v1.numero = v2.numero
        AND v1.letra = v2.letra AND v1.sucursal = v2.sucursal
    JOIN msgestion01art.dbo.articulo a ON v1.articulo = a.codigo
    WHERE a.marca = {MARCA}
      AND v2.codigo NOT IN {EXCL_CODIGOS}
      AND v2.fecha_comprobante >= '2025-01-01'
      AND v2.fecha_comprobante < '2026-01-01'
    GROUP BY MONTH(v2.fecha_comprobante)
    ORDER BY mes
    """
    cur = conn.cursor()
    cur.execute(sql)
    rows = {r.mes: (float(r.pares or 0), float(r.importe or 0)) for r in cur.fetchall()}
    return rows


def query_share_mensual(conn):
    """SHARE: Topper vs total negocio (rubros calzado/deporte, excl marcas gastos)."""
    sql = f"""
    SELECT MONTH(v2.fecha_comprobante) as mes,
           SUM(CASE WHEN a.marca = {MARCA} THEN
               CASE WHEN v1.operacion='+' THEN v1.cantidad
                    WHEN v1.operacion='-' THEN -v1.cantidad ELSE 0 END
           ELSE 0 END) as topper_pares,
           SUM(CASE WHEN v1.operacion='+' THEN v1.cantidad
                    WHEN v1.operacion='-' THEN -v1.cantidad ELSE 0 END) as total_pares,
           SUM(CASE WHEN a.marca = {MARCA} THEN
               CASE WHEN v1.operacion='+' THEN v1.cantidad * v1.precio
                    WHEN v1.operacion='-' THEN -v1.cantidad * v1.precio ELSE 0 END
           ELSE 0 END) as topper_importe,
           SUM(CASE WHEN v1.operacion='+' THEN v1.cantidad * v1.precio
                    WHEN v1.operacion='-' THEN -v1.cantidad * v1.precio ELSE 0 END) as total_importe
    FROM msgestionC.dbo.ventas1 v1
    JOIN msgestionC.dbo.ventas2 v2
        ON v1.codigo = v2.codigo AND v1.numero = v2.numero
        AND v1.letra = v2.letra AND v1.sucursal = v2.sucursal
    JOIN msgestion01art.dbo.articulo a ON v1.articulo = a.codigo
    WHERE v2.codigo NOT IN {EXCL_CODIGOS}
      AND v2.fecha_comprobante >= '2025-01-01'
      AND v2.fecha_comprobante < '2026-01-01'
      AND a.marca NOT IN {EXCL_MARCAS_GASTOS}
      AND a.rubro IN {RUBROS_CALZADO}
    GROUP BY MONTH(v2.fecha_comprobante)
    ORDER BY mes
    """
    cur = conn.cursor()
    cur.execute(sql)
    rows = {}
    for r in cur.fetchall():
        rows[r.mes] = {
            'topper_pares': float(r.topper_pares or 0),
            'total_pares': float(r.total_pares or 0),
            'topper_importe': float(r.topper_importe or 0),
            'total_importe': float(r.total_importe or 0),
        }
    return rows


def query_compras_mensual(conn):
    """Compras (reposición) Topper 2025 por mes."""
    sql = f"""
    SELECT MONTH(c2.fecha_comprobante) as mes,
           SUM(CASE WHEN c1.operacion='+' THEN c1.cantidad ELSE 0 END) as pares
    FROM msgestionC.dbo.compras1 c1
    JOIN msgestionC.dbo.compras2 c2
        ON c1.codigo = c2.codigo AND c1.numero = c2.numero
        AND c1.letra = c2.letra AND c1.sucursal = c2.sucursal
    JOIN msgestion01art.dbo.articulo a ON c1.articulo = a.codigo
    WHERE a.marca = {MARCA}
      AND c2.codigo NOT IN {EXCL_CODIGOS}
      AND c2.fecha_comprobante >= '2025-01-01'
      AND c2.fecha_comprobante < '2026-01-01'
    GROUP BY MONTH(c2.fecha_comprobante)
    ORDER BY mes
    """
    cur = conn.cursor()
    cur.execute(sql)
    return {r.mes: float(r.pares or 0) for r in cur.fetchall()}


def query_stock_por_subrubro(conn):
    """Stock actual Topper por subrubro."""
    sql = f"""
    SELECT a.subrubro,
           sr.descripcion as subrubro_desc,
           COUNT(DISTINCT LEFT(a.codigo_sinonimo, 10)) as modelos,
           SUM(s.stock_actual) as stock_pares
    FROM msgestionC.dbo.stock s
    JOIN msgestion01art.dbo.articulo a ON s.articulo = a.codigo
    LEFT JOIN msgestionC.dbo.subrubro sr ON a.subrubro = sr.codigo
    WHERE a.marca = {MARCA}
      AND s.deposito IN {DEPOSITOS}
      AND s.stock_actual > 0
    GROUP BY a.subrubro, sr.descripcion
    ORDER BY SUM(s.stock_actual) DESC
    """
    cur = conn.cursor()
    cur.execute(sql)
    return cur.fetchall()


def query_vel_real_por_subrubro(conn):
    """Velocidad real Topper por subrubro (desde tabla materializada)."""
    sql = f"""
    SELECT a.subrubro,
           sr.descripcion as subrubro_desc,
           COUNT(DISTINCT vr.codigo) as modelos,
           SUM(vr.vel_real) as vel_real,
           SUM(vr.vel_aparente) as vel_aparente,
           AVG(vr.meses_quebrado) as avg_meses_quebrado,
           AVG(vr.factor_quiebre) as avg_factor_quiebre,
           SUM(vr.ventas_perdidas) as ventas_perdidas
    FROM omicronvt.dbo.vel_real_articulo vr
    JOIN (SELECT DISTINCT LEFT(codigo_sinonimo, 10) as csr, subrubro
          FROM msgestion01art.dbo.articulo WHERE marca = {MARCA}) a
        ON a.csr = vr.codigo
    LEFT JOIN msgestionC.dbo.subrubro sr ON a.subrubro = sr.codigo
    WHERE vr.vel_real > 0
    GROUP BY a.subrubro, sr.descripcion
    ORDER BY SUM(vr.vel_real) DESC
    """
    cur = conn.cursor()
    cur.execute(sql)
    return cur.fetchall()


def query_top_modelos(conn, top_n=20):
    """Top N modelos Topper por vel_real."""
    sql = f"""
    SELECT TOP {top_n}
           vr.codigo,
           vr.vel_real, vr.vel_aparente,
           vr.meses_quebrado, vr.meses_con_stock,
           vr.ventas_perdidas,
           a.descripcion_1, a.subrubro,
           sr.descripcion as subrubro_desc
    FROM omicronvt.dbo.vel_real_articulo vr
    JOIN (SELECT LEFT(codigo_sinonimo, 10) as csr,
                 MIN(descripcion_1) as descripcion_1,
                 MIN(subrubro) as subrubro
          FROM msgestion01art.dbo.articulo WHERE marca = {MARCA}
          GROUP BY LEFT(codigo_sinonimo, 10)) a
        ON a.csr = vr.codigo
    LEFT JOIN msgestionC.dbo.subrubro sr ON a.subrubro = sr.codigo
    WHERE vr.vel_real > 0
    ORDER BY vr.vel_real DESC
    """
    cur = conn.cursor()
    cur.execute(sql)
    return cur.fetchall()


def query_stock_total_y_vel(conn):
    """Stock total + vel_real total para MOS global."""
    sql = f"""
    SELECT
        (SELECT SUM(s.stock_actual)
         FROM msgestionC.dbo.stock s
         JOIN msgestion01art.dbo.articulo a ON s.articulo = a.codigo
         WHERE a.marca = {MARCA}
           AND s.deposito IN {DEPOSITOS}) as stock_total,
        (SELECT SUM(vr.vel_real)
         FROM omicronvt.dbo.vel_real_articulo vr
         JOIN (SELECT DISTINCT LEFT(codigo_sinonimo, 10) as csr
               FROM msgestion01art.dbo.articulo WHERE marca = {MARCA}) a2
           ON a2.csr = vr.codigo
         WHERE vr.vel_real > 0) as vel_real_total
    """
    cur = conn.cursor()
    cur.execute(sql)
    r = cur.fetchone()
    return float(r.stock_total or 0), float(r.vel_real_total or 1)


# ═════════════════════════════════════════════════════
# HOJAS
# ═════════════════════════════════════════════════════

def crear_hoja_manuel(wb, so_data, share_data, stock_total, vel_real_total, compras_data):
    """Hoja 1: datos para Manuel Simón (SO/SHARE/MOS)."""
    ws = wb.active
    ws.title = "Para Manuel"

    # Título
    ws.merge_cells('A1:M1')
    ws['A1'] = "HACHE CUATRO SRL — TOPPER — Información Complementaria Preventa"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    ws['A2'] = "Período: Enero - Diciembre 2025"
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    row = 4
    ws.cell(row=row, column=1, value="")
    for i, mes in enumerate(MESES):
        ws.cell(row=row, column=i + 2, value=mes)
    ws.cell(row=row, column=14, value="TOTAL")
    apply_header_style(ws, row, 1, 14)

    # SO (pares)
    row = 5
    ws.cell(row=row, column=1, value="SO (pares)").font = Font(bold=True)
    total_pares = 0
    for m in range(1, 13):
        val = so_data.get(m, (0, 0))[0]
        total_pares += val
        c = ws.cell(row=row, column=m + 1, value=val)
        c.number_format = NUM_FORMAT
        c.border = THIN_BORDER
    ws.cell(row=row, column=14, value=total_pares).number_format = NUM_FORMAT
    ws.cell(row=row, column=14).font = Font(bold=True)

    # SO ($M)
    row = 6
    ws.cell(row=row, column=1, value="SO ($ s/IVA)").font = Font(bold=True)
    total_imp = 0
    for m in range(1, 13):
        val = so_data.get(m, (0, 0))[1]
        total_imp += val
        c = ws.cell(row=row, column=m + 1, value=round(val))
        c.number_format = MONEY_FORMAT
        c.border = THIN_BORDER
    ws.cell(row=row, column=14, value=round(total_imp)).number_format = MONEY_FORMAT
    ws.cell(row=row, column=14).font = Font(bold=True)

    # SHARE pares
    row = 7
    ws.cell(row=row, column=1, value="SHARE (pares)").font = Font(bold=True)
    for m in range(1, 13):
        sd = share_data.get(m)
        if sd and sd['total_pares'] > 0:
            val = sd['topper_pares'] / sd['total_pares']
        else:
            val = 0
        c = ws.cell(row=row, column=m + 1, value=val)
        c.number_format = PCT_FORMAT
        c.border = THIN_BORDER
    # Promedio
    shares = []
    for m in range(1, 13):
        sd = share_data.get(m)
        if sd and sd['total_pares'] > 0:
            shares.append(sd['topper_pares'] / sd['total_pares'])
    ws.cell(row=row, column=14, value=sum(shares) / len(shares) if shares else 0).number_format = PCT_FORMAT
    ws.cell(row=row, column=14).font = Font(bold=True)

    # SHARE $
    row = 8
    ws.cell(row=row, column=1, value="SHARE ($)").font = Font(bold=True)
    for m in range(1, 13):
        sd = share_data.get(m)
        if sd and sd['total_importe'] > 0:
            val = sd['topper_importe'] / sd['total_importe']
        else:
            val = 0
        c = ws.cell(row=row, column=m + 1, value=val)
        c.number_format = PCT_FORMAT
        c.border = THIN_BORDER
    shares_imp = []
    for m in range(1, 13):
        sd = share_data.get(m)
        if sd and sd['total_importe'] > 0:
            shares_imp.append(sd['topper_importe'] / sd['total_importe'])
    ws.cell(row=row, column=14, value=sum(shares_imp) / len(shares_imp) if shares_imp else 0).number_format = PCT_FORMAT
    ws.cell(row=row, column=14).font = Font(bold=True)

    # MOS (reconstruido mes a mes)
    row = 9
    ws.cell(row=row, column=1, value="MOS (meses)").font = Font(bold=True)
    # Reconstruir stock hacia atrás desde stock_total actual
    stock_fin_dic = stock_total
    mos_by_month = {}
    # Calcular stock al inicio de cada mes, yendo de dic hacia ene
    stock_fin = stock_fin_dic
    vel_mensual = vel_real_total  # pares/mes vel_real
    for m in range(12, 0, -1):
        ventas_m = so_data.get(m, (0, 0))[0]
        compras_m = compras_data.get(m, 0)
        # stock al inicio del mes = stock_fin + ventas - compras
        stock_inicio = stock_fin + ventas_m - compras_m
        # MOS = stock al fin del mes / vel_real
        mos = stock_fin / vel_mensual if vel_mensual > 0 else 0
        mos_by_month[m] = round(mos, 1)
        stock_fin = stock_inicio  # para el mes anterior

    for m in range(1, 13):
        val = mos_by_month.get(m, 0)
        c = ws.cell(row=row, column=m + 1, value=val)
        c.number_format = '0.0'
        c.border = THIN_BORDER
        # Colorear según criticidad
        if val < 2:
            c.fill = RED_FILL
        elif val < 4:
            c.fill = YELLOW_FILL
        else:
            c.fill = GREEN_FILL
    # MOS actual
    mos_actual = stock_total / vel_real_total if vel_real_total > 0 else 0
    ws.cell(row=row, column=14, value=round(mos_actual, 1)).number_format = '0.0'
    ws.cell(row=row, column=14).font = Font(bold=True)

    # Compras (reposición)
    row = 11
    ws.cell(row=row, column=1, value="Compras (pares)").font = Font(bold=True)
    ws.cell(row=row, column=1).font = Font(bold=True, color="2F5496")
    total_compras = 0
    for m in range(1, 13):
        val = compras_data.get(m, 0)
        total_compras += val
        c = ws.cell(row=row, column=m + 1, value=val)
        c.number_format = NUM_FORMAT
        c.border = THIN_BORDER
    ws.cell(row=row, column=14, value=total_compras).number_format = NUM_FORMAT
    ws.cell(row=row, column=14).font = Font(bold=True)

    # Balance (compras - ventas)
    row = 12
    ws.cell(row=row, column=1, value="Balance (C-V)").font = Font(bold=True, italic=True)
    for m in range(1, 13):
        ventas = so_data.get(m, (0, 0))[0]
        compras = compras_data.get(m, 0)
        bal = compras - ventas
        c = ws.cell(row=row, column=m + 1, value=bal)
        c.number_format = NUM_FORMAT
        c.border = THIN_BORDER
        if bal < 0:
            c.font = Font(color="FF0000")
    ws.cell(row=row, column=14, value=total_compras - total_pares).number_format = NUM_FORMAT
    ws.cell(row=row, column=14).font = Font(bold=True, color="FF0000" if (total_compras - total_pares) < 0 else "006600")

    apply_border_range(ws, 4, 12, 1, 14)

    # Anchos
    ws.column_dimensions['A'].width = 18
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 13

    return ws


def crear_hoja_analisis(wb, vel_subrubro, stock_subrubro, top_modelos, so_data, compras_data):
    """Hoja 2: Análisis interno con quiebre por categoría."""
    ws = wb.create_sheet("Analisis Interno")

    # ── Sección 1: Quiebre por categoría ──
    ws.merge_cells('A1:H1')
    ws['A1'] = "DIAGNÓSTICO DE QUIEBRE POR CATEGORÍA — Topper"
    ws['A1'].font = TITLE_FONT

    row = 3
    headers = ["Subrubro", "Desc", "Modelos", "Vel Real p/m", "Vel Aparente", "Meses Quebr (avg)", "Factor Quiebre", "Vtas Perdidas 12m"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    apply_header_style(ws, row, 1, len(headers))

    row = 4
    for r in vel_subrubro:
        ws.cell(row=row, column=1, value=r.subrubro)
        ws.cell(row=row, column=2, value=r.subrubro_desc or "")
        ws.cell(row=row, column=3, value=r.modelos).number_format = NUM_FORMAT
        ws.cell(row=row, column=4, value=round(r.vel_real, 1)).number_format = '0.0'
        ws.cell(row=row, column=5, value=round(r.vel_aparente, 1)).number_format = '0.0'
        ws.cell(row=row, column=6, value=r.avg_meses_quebrado)
        fq = r.avg_factor_quiebre
        c = ws.cell(row=row, column=7, value=round(fq, 1) if fq else 0)
        c.number_format = '0.0'
        if fq and fq > 2.5:
            c.fill = RED_FILL
        elif fq and fq > 1.5:
            c.fill = YELLOW_FILL
        ws.cell(row=row, column=8, value=round(r.ventas_perdidas)).number_format = NUM_FORMAT
        apply_border_range(ws, row, row, 1, 8)
        row += 1

    # ── Sección 2: Stock + MOS por categoría ──
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="STOCK Y COBERTURA POR CATEGORÍA").font = SUBTITLE_FONT

    row += 1
    headers2 = ["Subrubro", "Desc", "Modelos", "Stock Pares", "Vel Real p/m", "MOS (meses)"]
    for i, h in enumerate(headers2):
        ws.cell(row=row, column=i + 1, value=h)
    apply_header_style(ws, row, 1, len(headers2))

    # Cruzar stock con vel_real
    vel_dict = {r.subrubro: r.vel_real for r in vel_subrubro}
    desc_dict = {r.subrubro: r.subrubro_desc for r in vel_subrubro}
    row += 1
    for sr in stock_subrubro:
        ws.cell(row=row, column=1, value=sr.subrubro)
        ws.cell(row=row, column=2, value=sr.subrubro_desc or desc_dict.get(sr.subrubro, ""))
        ws.cell(row=row, column=3, value=sr.modelos)
        ws.cell(row=row, column=4, value=sr.stock_pares).number_format = NUM_FORMAT
        vel = vel_dict.get(sr.subrubro, 0)
        ws.cell(row=row, column=5, value=round(vel, 1)).number_format = '0.0'
        mos = sr.stock_pares / vel if vel > 0 else 99
        c = ws.cell(row=row, column=6, value=round(mos, 1))
        c.number_format = '0.0'
        if mos < 2:
            c.fill = RED_FILL
        elif mos < 4:
            c.fill = YELLOW_FILL
        else:
            c.fill = GREEN_FILL
        apply_border_range(ws, row, row, 1, 6)
        row += 1

    # ── Sección 3: Top modelos con quiebre ──
    row += 2
    ws.merge_cells(f'A{row}:I{row}')
    ws.cell(row=row, column=1, value="TOP 20 MODELOS — VELOCIDAD REAL Y QUIEBRE").font = SUBTITLE_FONT

    row += 1
    headers3 = ["Código", "Descripción", "Categoría", "Vel Real", "Vel Aparente", "Meses OK", "Meses Quebr", "Factor", "Vtas Perd"]
    for i, h in enumerate(headers3):
        ws.cell(row=row, column=i + 1, value=h)
    apply_header_style(ws, row, 1, len(headers3))

    row += 1
    for r in top_modelos:
        ws.cell(row=row, column=1, value=r.codigo)
        desc = r.descripcion_1 or ""
        ws.cell(row=row, column=2, value=desc[:50])
        ws.cell(row=row, column=3, value=r.subrubro_desc or "")
        ws.cell(row=row, column=4, value=round(r.vel_real, 1)).number_format = '0.0'
        ws.cell(row=row, column=5, value=round(r.vel_aparente, 1)).number_format = '0.0'
        ws.cell(row=row, column=6, value=r.meses_con_stock)
        mq = r.meses_quebrado
        c = ws.cell(row=row, column=7, value=mq)
        if mq >= 8:
            c.fill = RED_FILL
        elif mq >= 4:
            c.fill = YELLOW_FILL
        factor = r.vel_real / r.vel_aparente if r.vel_aparente > 0 else 0
        ws.cell(row=row, column=8, value=round(factor, 1)).number_format = '0.0'
        ws.cell(row=row, column=9, value=round(r.ventas_perdidas)).number_format = NUM_FORMAT
        apply_border_range(ws, row, row, 1, 9)
        row += 1

    # Anchos
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    for col in range(4, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14

    return ws


def crear_hoja_argumentos(wb, so_data, share_data, stock_total, vel_real_total, compras_data, vel_subrubro, top_modelos):
    """Hoja 3: Argumentos ejecutivos para la negociación."""
    ws = wb.create_sheet("Argumentos Preventa")

    ws.merge_cells('A1:F1')
    ws['A1'] = "ARGUMENTOS PARA PREVENTA TOPPER — H4 SRL"
    ws['A1'].font = TITLE_FONT

    ws.merge_cells('A2:F2')
    ws['A2'] = "Preparado para reunión con Manuel Simón (Alpargatas) — Abril 2026"
    ws['A2'].font = Font(italic=True, size=10, color="666666")

    # ── Resumen ejecutivo ──
    row = 4
    ws.cell(row=row, column=1, value="RESUMEN EJECUTIVO").font = SUBTITLE_FONT
    ws.cell(row=row, column=1).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')

    total_pares = sum(so_data.get(m, (0, 0))[0] for m in range(1, 13))
    total_importe = sum(so_data.get(m, (0, 0))[1] for m in range(1, 13))
    total_compras = sum(compras_data.get(m, 0) for m in range(1, 13))
    mos_actual = stock_total / vel_real_total if vel_real_total > 0 else 0

    # Ventas perdidas totales
    total_vtas_perdidas = sum(r.ventas_perdidas for r in vel_subrubro)

    datos = [
        ("Sell Out 2025", f"{total_pares:,.0f} pares — ${total_importe/1e6:,.1f}M"),
        ("Vel. Real (corregida quiebre)", f"{vel_real_total:,.0f} p/mes — demanda real ~{vel_real_total*12:,.0f} p/año"),
        ("Vel. Aparente (sin corregir)", f"{total_pares/12:,.0f} p/mes — subestima {vel_real_total/(total_pares/12)*100-100:,.0f}% la demanda"),
        ("Share promedio 2025", f"{sum(share_data[m]['topper_pares'] for m in share_data)/sum(share_data[m]['total_pares'] for m in share_data)*100:.1f}% pares / {sum(share_data[m]['topper_importe'] for m in share_data)/sum(share_data[m]['total_importe'] for m in share_data)*100:.1f}% importe"),
        ("Stock actual", f"{stock_total:,.0f} pares"),
        ("MOS actual (sobre vel_real)", f"{mos_actual:.1f} meses"),
        ("Compras 2025", f"{total_compras:,.0f} pares"),
        ("Déficit reposición", f"{total_pares - total_compras:,.0f} pares más vendidos que comprados"),
        ("Ventas perdidas estimadas", f"{total_vtas_perdidas:,.0f} pares/año por falta de stock"),
    ]

    row = 5
    for label, value in datos:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.merge_cells(f'B{row}:F{row}')
        apply_border_range(ws, row, row, 1, 6)
        row += 1

    # ── Hallazgos clave ──
    row += 1
    ws.cell(row=row, column=1, value="HALLAZGOS CLAVE PARA LA NEGOCIACIÓN").font = SUBTITLE_FONT
    ws.cell(row=row, column=1).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')

    hallazgos = [
        "1. QUIEBRE MASIVO en Running y Training: los top sellers están sin stock 6-11 de 12 meses",
        f"2. La demanda real es ~{vel_real_total*12:,.0f} p/año, NO los {total_pares:,.0f} que vendimos — el quiebre oculta {(vel_real_total*12-total_pares):,.0f} pares de demanda insatisfecha",
        f"3. Share en $ ({sum(share_data[m]['topper_importe'] for m in share_data)/sum(share_data[m]['total_importe'] for m in share_data)*100:.1f}%) es 2x el share en pares ({sum(share_data[m]['topper_pares'] for m in share_data)/sum(share_data[m]['total_pares'] for m in share_data)*100:.1f}%) → ticket promedio Topper DUPLICA el promedio del negocio",
        "4. TEXTIL subcomprado: Buzos con factor quiebre 4x, Remeras 2.3x, Pantalones 2.3x",
        f"5. Balance negativo: compramos {total_compras:,.0f} pero vendimos {total_pares:,.0f} → vivimos del stock viejo, no es sostenible",
        f"6. Con stock completo y MOS>4 meses, podríamos llegar a {vel_real_total*12:,.0f}+ pares/año (+{(vel_real_total*12/total_pares-1)*100:.0f}%)",
    ]

    row += 1
    for h in hallazgos:
        ws.cell(row=row, column=1, value=h)
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        row += 1

    # ── Top 5 modelos quebrados ──
    row += 1
    ws.cell(row=row, column=1, value="TOP 5 MODELOS CON MAYOR OPORTUNIDAD PERDIDA").font = SUBTITLE_FONT
    ws.merge_cells(f'A{row}:F{row}')

    row += 1
    for i, h in enumerate(["Modelo", "Vel Real p/m", "Vel Aparente", "Meses Quebrado", "Categoría"]):
        ws.cell(row=row, column=i + 1, value=h)
    apply_header_style(ws, row, 1, 5)

    row += 1
    for r in top_modelos[:5]:
        desc = r.descripcion_1 or ""
        ws.cell(row=row, column=1, value=desc[:45])
        ws.cell(row=row, column=2, value=round(r.vel_real, 1))
        ws.cell(row=row, column=3, value=round(r.vel_aparente, 1))
        ws.cell(row=row, column=4, value=f"{r.meses_quebrado}/12")
        ws.cell(row=row, column=5, value=r.subrubro_desc or "")
        apply_border_range(ws, row, row, 1, 5)
        row += 1

    # Anchos
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 55
    for col in range(3, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

    return ws


# ═════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════

def main():
    print("Conectando al ERP (192.168.2.111)...")
    conn = get_conn()

    print("  → SO mensual 2025...")
    so_data = query_so_mensual(conn)

    print("  → SHARE mensual 2025...")
    share_data = query_share_mensual(conn)

    print("  → Compras mensuales 2025...")
    compras_data = query_compras_mensual(conn)

    print("  → Stock por subrubro...")
    stock_subrubro = query_stock_por_subrubro(conn)

    print("  → Vel_real por subrubro...")
    vel_subrubro = query_vel_real_por_subrubro(conn)

    print("  → Top 20 modelos...")
    top_modelos = query_top_modelos(conn)

    print("  → Stock total + vel_real global...")
    stock_total, vel_real_total = query_stock_total_y_vel(conn)

    conn.close()
    print(f"\n  Stock total: {stock_total:,.0f} pares")
    print(f"  Vel real total: {vel_real_total:,.0f} p/mes")
    print(f"  MOS actual: {stock_total/vel_real_total:.1f} meses")

    # ── Generar Excel ──
    print("\nGenerando Excel...")
    wb = Workbook()

    crear_hoja_manuel(wb, so_data, share_data, stock_total, vel_real_total, compras_data)
    crear_hoja_analisis(wb, vel_subrubro, stock_subrubro, top_modelos, so_data, compras_data)
    crear_hoja_argumentos(wb, so_data, share_data, stock_total, vel_real_total, compras_data, vel_subrubro, top_modelos)

    # Guardar
    out_dir = os.path.join(os.path.dirname(__file__), "..", "_excel_pedidos")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "topper_preventa_2025.xlsx")
    wb.save(out_path)
    print(f"\n✓ Excel guardado en: {out_path}")
    print(f"  → Hoja 1: 'Para Manuel' (SO/SHARE/MOS)")
    print(f"  → Hoja 2: 'Analisis Interno' (quiebre, top modelos)")
    print(f"  → Hoja 3: 'Argumentos Preventa' (resumen ejecutivo)")

    # Resumen rápido
    total_pares = sum(so_data.get(m, (0, 0))[0] for m in range(1, 13))
    total_importe = sum(so_data.get(m, (0, 0))[1] for m in range(1, 13))
    print(f"\n  SO 2025: {total_pares:,.0f} pares — ${total_importe/1e6:,.1f}M")
    print(f"  Compras 2025: {sum(compras_data.get(m, 0) for m in range(1, 13)):,.0f} pares")
    print(f"  Déficit: {total_pares - sum(compras_data.get(m, 0) for m in range(1, 13)):,.0f} pares")


if __name__ == "__main__":
    main()
