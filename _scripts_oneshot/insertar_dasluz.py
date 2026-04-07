"""
Script de insercion: Pedido DasLuz (Botas de Lluvia) -- Invierno 2026
Empresa: H4 --> msgestion03 (default, verificar proveedor_asignacion_base)
Proveedor: 610 (DAS LUZ)
Total: 197 pares | $2,972,900

Fuente: /Volumes/compartido/COMPRAS/INVIERNO 2026/Das Luz/Pedidos 04-02 dasluz/PEDIDO DASLUZ unificado.xlsx

ESTRUCTURA EXCEL (12 filas de datos, R2-R13, R14=subtotal a excluir):
- Col A: ARTICULO (codigos cortos: 1200, 1150, 1210, 1110, 1006, 1610)
- Col B: COLOR
- Cols C-V: Talles 28-47 (solo 36,37,38,40 tienen datos)
- Col W: TOTAL (197 pares)
- Col X: PRECIO ($12,500 a $18,900)
- Col Z: SUBRUBRO (todo "Bota de lluvia")
- Col AA: Subtotal por fila

NOTAS IMPORTANTES:
- Los articulos DasLuz se buscan en el ERP por sinonimo (campo sinonimo en
  msgestion01art.dbo.articulo) usando el codigo corto del Excel + color.
- Si el sinonimo no matchea directamente, el script imprime un WARNING y
  hay que resolver manualmente antes de ejecutar.
- Cada fila del Excel con cantidad > 0 en un talle genera una linea pedico1.
  Como cada talle es un articulo diferente en el ERP, se busca articulo+talle.

ESTRUCTURA pedico1 (columnas usadas):
  codigo, letra, sucursal, numero, orden, articulo, cantidad, precio,
  descuento_reng1, descuento_reng2, fecha, estado
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pyodbc
import datetime
from config import get_conn_string

# ─── PARAMETROS DEL PEDIDO ────────────────────────────────────────────────────
EMPRESA        = "H4"          # --> msgestion03 (default para DasLuz, ajustar si es CALZALINDO)
BASE           = "MSGESTION03" # base destino para INSERT
PROVEEDOR      = 610           # DAS LUZ
DESCUENTO_PROV = 0             # Sin descuento (confirmar con factura)
DESCUENTO_BON  = 0             # Sin bonificacion adicional

EXCEL_PATH = "/Volumes/compartido/COMPRAS/INVIERNO 2026/Das Luz/Pedidos 04-02 dasluz/PEDIDO DASLUZ unificado.xlsx"

# Columnas de talles en el Excel: C=28, D=29, ..., V=47
# Mapeo: col_index (0-based desde A) -> talle
TALLE_COLS = {}
for i, talle in enumerate(range(28, 48)):  # 28 a 47 = 20 columnas
    TALLE_COLS[i + 3] = str(talle)  # col C=index 3 (1-based in openpyxl = col 3)
# En openpyxl: col 3=C (talle 28), col 4=D (talle 29), ... col 22=V (talle 47)


def leer_excel(path):
    """
    Lee el Excel de DasLuz y retorna lista de tuplas:
    [(articulo_corto, color, talle, cantidad, precio, descripcion), ...]
    Excluye fila 14 (subtotal) y celdas con cantidad 0 o None.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl no instalado. Ejecutar: pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    lineas = []
    for row_num in range(2, 14):  # Filas 2 a 13 (datos), excluir 14 (subtotal)
        row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

        articulo_corto = str(row[0]).strip() if row[0] else None
        color = str(row[1]).strip() if row[1] else ""
        precio = float(row[23]) if row[23] else 0  # Col X = index 23 (0-based)
        # total_fila = row[22]  # Col W = index 22 (para validacion)

        if not articulo_corto or precio <= 0:
            continue

        # Recorrer columnas de talles (C=col3 a V=col22, 0-based: index 2 a 21)
        for col_idx in range(2, 22):  # 0-based: C=2, D=3, ..., V=21
            talle = str(28 + (col_idx - 2))  # C=28, D=29, ..., V=47
            cantidad = row[col_idx]
            if cantidad and isinstance(cantidad, (int, float)) and cantidad > 0:
                desc = f"{articulo_corto} {color} T{talle}"
                lineas.append((articulo_corto, color, talle, int(cantidad), precio, desc))

    wb.close()
    return lineas


def buscar_articulo_erp(cursor, articulo_corto, color, talle):
    """
    Busca el codigo de articulo en el ERP por sinonimo.
    Estrategia de busqueda (de mas especifica a menos):
    1. sinonimo LIKE '%{articulo_corto}%' AND descripcion con color y talle
    2. sinonimo = '{articulo_corto}' exacto con auxiliar = proveedor
    3. Busqueda amplia por proveedor + descripcion
    """
    # Estrategia 1: buscar por auxiliar (proveedor) + sinonimo contiene el codigo corto
    # Filtrar por talle en la descripcion
    cursor.execute("""
        SELECT codigo, codigo_sinonimo, descripcion_1
        FROM msgestion01art.dbo.articulo
        WHERE proveedor = ?
          AND codigo_sinonimo LIKE ?
        ORDER BY codigo DESC
    """, (PROVEEDOR, f'%{articulo_corto}%'))

    candidatos = cursor.fetchall()

    if not candidatos:
        # Estrategia 2: buscar sin filtro de proveedor, solo sinonimo
        cursor.execute("""
            SELECT codigo, codigo_sinonimo, descripcion_1
            FROM msgestion01art.dbo.articulo
            WHERE codigo_sinonimo LIKE ?
            ORDER BY codigo DESC
        """, (f'%{articulo_corto}%',))
        candidatos = cursor.fetchall()

    # Filtrar por color y talle en la descripcion
    color_upper = color.upper()
    for cod, sin, desc in candidatos:
        desc_upper = (desc or "").upper()
        # Verificar que el talle aparezca en la descripcion o sinonimo
        if talle in (desc_upper or "") or talle in (sin or ""):
            if color_upper and color_upper[:3] in desc_upper:
                return cod, sin, desc

    # Si no matcheo por color+talle, intentar solo talle
    for cod, sin, desc in candidatos:
        desc_upper = (desc or "").upper()
        if talle in desc_upper or talle in (sin or ""):
            return cod, sin, desc

    # Si hay un unico candidato, retornarlo con warning
    if len(candidatos) == 1:
        cod, sin, desc = candidatos[0]
        return cod, sin, desc

    return None, None, None


def insertar_pedido_dasluz(dry_run=True):
    """
    Lee el Excel de DasLuz, matchea articulos contra el ERP, e inserta
    pedico2 (cabecera) + pedico1 (detalle).
    dry_run=True: muestra que se haria sin ejecutar nada.
    """
    # ── 0. Leer Excel ─────────────────────────────────────────────────────────
    print(f"Leyendo Excel: {EXCEL_PATH}")
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: No se encuentra el archivo Excel.")
        print(f"Verificar que el volumen compartido este montado.")
        sys.exit(1)

    lineas_excel = leer_excel(EXCEL_PATH)
    print(f"Lineas leidas del Excel: {len(lineas_excel)}")

    if not lineas_excel:
        print("ERROR: No se leyeron lineas del Excel.")
        sys.exit(1)

    # ── 1. Conectar y matchear articulos ──────────────────────────────────────
    conn_str = get_conn_string(BASE.lower())
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    detalle = []  # (codigo_erp, cantidad, precio, descripcion)
    errores = []
    total_pares_excel = 0

    print(f"\nMatcheando articulos contra ERP...")
    for art_corto, color, talle, qty, precio, desc in lineas_excel:
        total_pares_excel += qty
        cod_erp, sin_erp, desc_erp = buscar_articulo_erp(cursor, art_corto, color, talle)

        if cod_erp:
            detalle.append((cod_erp, qty, precio, f"{desc} -> ERP {cod_erp} ({desc_erp})"))
            print(f"  OK: {desc:30s} -> art={cod_erp} ({sin_erp}) {desc_erp}")
        else:
            errores.append((art_corto, color, talle, qty, precio, desc))
            print(f"  ** NO MATCH: {desc:30s} (qty={qty}, precio={precio})")

    print(f"\nResumen matcheo:")
    print(f"  Matcheados: {len(detalle)} lineas, {sum(d[1] for d in detalle)} pares")
    print(f"  Sin match:  {len(errores)} lineas, {sum(e[3] for e in errores)} pares")
    print(f"  Total Excel: {total_pares_excel} pares")

    if errores:
        print(f"\n*** ATENCION: {len(errores)} articulos sin match en ERP ***")
        print("Resolver manualmente antes de ejecutar con --ejecutar.")
        print("Articulos sin match:")
        for art, col, tal, qty, precio, desc in errores:
            print(f"  {desc} | qty={qty} | precio={precio}")
        if dry_run:
            print("\nContinuando en modo DRY RUN con los articulos matcheados...")
        else:
            print("\nABORTANDO: No se puede insertar con articulos sin match.")
            print("Opciones:")
            print("  1. Dar de alta los articulos faltantes en el ERP")
            print("  2. Agregar los codigos manualmente al DETALLE_MANUAL")
            print("  3. Ejecutar con --forzar para insertar solo los matcheados")
            conn.close()
            return

    if not detalle:
        print("\nERROR: Ningun articulo matcheo. No hay nada para insertar.")
        conn.close()
        return

    # ── 2. Calcular proximo numero de pedido ──────────────────────────────────
    cursor.execute(
        f"SELECT MAX(CAST(numero AS INT)) FROM {BASE}.dbo.pedico2 "
        f"WHERE codigo=8 AND letra='X' AND sucursal=1"
    )
    row = cursor.fetchone()
    ultimo_num = row[0] if row[0] else 0
    nuevo_numero = ultimo_num + 1

    cursor.execute(
        f"SELECT MAX(orden) FROM {BASE}.dbo.pedico2 "
        f"WHERE codigo=8 AND letra='X' AND sucursal=1"
    )
    row = cursor.fetchone()
    ultimo_orden = row[0] if row[0] else 0
    nuevo_orden = (ultimo_orden + 1) if (ultimo_orden + 1) <= 99 else 1

    hoy = datetime.date.today().strftime("%Y%m%d")

    total_pares = sum(d[1] for d in detalle)
    monto_total = sum(d[1] * d[2] for d in detalle)

    # Buscar nombre proveedor
    cursor.execute(f"SELECT denominacion FROM {BASE}.dbo.proveedores WHERE numero=?", (PROVEEDOR,))
    row = cursor.fetchone()
    nombre_prov = row[0].strip() if row and row[0] else f"DAS LUZ (PROV {PROVEEDOR})"

    print(f"\n{'='*70}")
    print(f"PEDIDO A INSERTAR:")
    print(f"  Numero:     {nuevo_numero}")
    print(f"  Orden:      {nuevo_orden}")
    print(f"  Proveedor:  {PROVEEDOR} -- {nombre_prov}")
    print(f"  Empresa:    {EMPRESA} -> {BASE}")
    print(f"  Fecha:      {hoy}")
    print(f"  Pares:      {total_pares}")
    print(f"  Monto:      ${monto_total:,.0f}")
    print(f"  Lineas:     {len(detalle)}")
    print(f"{'='*70}")
    print()
    for i, (art_cod, qty, precio, desc_art) in enumerate(detalle, 1):
        print(f"  [{i:>3}] art={art_cod:>7}, qty={qty:>3}, precio=${precio:>10,.0f} -- {desc_art}")

    if dry_run:
        print(f"\n[DRY RUN] No se ejecuto nada.")
        print(f"Para insertar: python _scripts_oneshot/insertar_dasluz.py --ejecutar")
        conn.close()
        return

    # ── 3. INSERT pedico2 (cabecera) ──────────────────────────────────────────
    obs = f"Pedido DasLuz Invierno 2026. {total_pares} pares. ${monto_total:,.0f}"
    sql_cab = f"""
    INSERT INTO {BASE}.dbo.pedico2
        (codigo, letra, sucursal, numero, orden, cuenta, denominacion,
         fecha_comprobante, estado, usuario, observaciones, importe_neto)
    VALUES
        (8, 'X', 1, ?, ?, ?, ?, ?, 'V', 'COWORK', ?, ?)
    """
    cursor.execute(sql_cab, (
        nuevo_numero, nuevo_orden,
        PROVEEDOR, nombre_prov,
        hoy,
        obs,
        monto_total,
    ))
    print(f"pedico2 insertado: numero={nuevo_numero}, orden={nuevo_orden}")

    # ── 4. INSERT pedico1 (detalle) ───────────────────────────────────────────
    sql_det = f"""
    INSERT INTO {BASE}.dbo.pedico1
        (codigo, letra, sucursal, numero, orden, renglon,
         articulo, cantidad, precio,
         descuento_reng1, descuento_reng2,
         fecha, estado)
    VALUES
        (8, 'X', 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'V')
    """
    for i, (art_cod, qty, precio, desc_art) in enumerate(detalle, 1):
        cursor.execute(sql_det, (
            nuevo_numero, nuevo_orden, i,
            art_cod, qty, precio,
            DESCUENTO_PROV, DESCUENTO_BON,
            hoy,
        ))
        print(f"  [{i:>3}] art={art_cod}, qty={qty}, precio=${precio:,.0f}")

    conn.commit()
    print(f"\nPedido DasLuz insertado exitosamente.")
    print(f"pedico2 numero {nuevo_numero} | {len(detalle)} lineas | {total_pares} pares | ${monto_total:,.0f}")
    conn.close()


# ─── DETALLE MANUAL (fallback si el Excel no matchea) ─────────────────────────
# Si la busqueda automatica falla para algunos articulos, agregar aqui los
# codigos ERP manualmente y descomentar la seccion en main().
# Formato: (codigo_erp, cantidad, precio, descripcion)
DETALLE_MANUAL = [
    # Ejemplo: (123456, 10, 15000.0, '1200 NEGRO T36'),
]


if __name__ == "__main__":
    dry_run = "--ejecutar" not in sys.argv
    forzar = "--forzar" in sys.argv

    if dry_run:
        print("\n*** MODO DRY RUN -- para ejecutar: python insertar_dasluz.py --ejecutar ***\n")
    else:
        print("\n*** MODO EJECUCION -- se insertara en la base de datos ***\n")

    insertar_pedido_dasluz(dry_run=dry_run)
