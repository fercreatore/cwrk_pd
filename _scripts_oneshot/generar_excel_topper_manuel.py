"""
generar_excel_topper_manuel.py
==============================
Versión EXTERNA para Manuel Simón (Alpargatas) — Preventa VERANO 2025/26.

Foco: temporada PV (verano), pico colegial diciembre.
Muestra: SO pares | SHARE pares | MOS mensual — nada más.

NO incluye: share $, márgenes, quiebre interno, vel_real, compras.

Ejecutar: python _scripts_oneshot/generar_excel_topper_manuel.py
Output:   _excel_pedidos/topper_preventa_verano2025_MANUEL.xlsx
"""

import sys, os
from decimal import Decimal
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import config
import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

MARCA = 314
EXCL_CODIGOS = "(7, 36)"
EXCL_MARCAS_GASTOS = "(1316, 1317, 1158, 436)"
RUBROS_CALZADO = "(1, 3, 4, 5, 6)"
DEPOSITOS = "(0,1,2,3,4,5,6,7,8,9,11,12,14,15,198)"

# Cono Sur — calendario comercial real:
#   Colegial    = FEBRERO (vuelta al cole)
#   Día del Niño = 3° domingo de AGOSTO
#   Día de la Madre = 3° domingo de OCTUBRE
#   Temporada verano (PV): entregas Jul-Dic, pico de ventas Ene-Mar
MESES_LABEL = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
               "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
# Meses de la temporada verano (entregas Jul-Dic, venta Oct-Mar)
MESES_VERANO = {7, 8, 9, 10, 11, 12}
MESES_COLEGIAL = {2}       # FEBRERO: pico colegial (vuelta al cole cono sur)
MESES_DIA_NINO = {8}       # AGOSTO: Día del Niño
MESES_DIA_MADRE = {10}     # OCTUBRE: Día de la Madre

# ── Estilos ──────────────────────────────────────────
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL   = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
VERANO_FILL   = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")  # azul claro
COLEGIAL_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # dorado
TOTAL_FILL    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # verde claro
RED_FILL      = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
YELLOW_FILL   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GREEN_FILL    = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
TITLE_FONT    = Font(bold=True, size=14, color="2F5496")
BOLD          = Font(bold=True)
THIN_BORDER   = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)
PCT_FORMAT    = '0.0%'
NUM_FORMAT    = '#,##0'


def get_conn():
    return pyodbc.connect(config.CONN_COMPRAS)


# ═════════════════════════════════════════════════════
# QUERIES
# ═════════════════════════════════════════════════════

def query_so_mensual_anio(conn, anio):
    """SO Topper mensual para un año dado.
    Usa omicron_ventas1 + v.fecha (igual que el ranking web2py).
    NO usa ventas2.fecha_comprobante (fecha contable, difiere de fecha real de venta).
    """
    sql = f"""
    SELECT MONTH(v.fecha) as mes,
           SUM(CASE WHEN v.operacion='+' THEN v.cantidad
                    WHEN v.operacion='-' THEN -v.cantidad ELSE 0 END) as pares
    FROM msgestion01.dbo.omicron_ventas1 v
    LEFT JOIN msgestion01art.dbo.articulo a ON v.articulo = a.codigo
    WHERE a.marca = {MARCA}
      AND v.codigo NOT IN {EXCL_CODIGOS}
      AND v.fecha >= '{anio}-01-01'
      AND v.fecha <= '{anio}-12-31'
    GROUP BY MONTH(v.fecha)
    ORDER BY mes
    """
    cur = conn.cursor()
    cur.execute(sql)
    return {r.mes: float(r.pares or 0) for r in cur.fetchall()}


def query_share_mensual_anio(conn, anio):
    """SHARE Topper pares vs total negocio. Usa omicron_ventas1 + v.fecha (consistente con SO)."""
    sql = f"""
    SELECT MONTH(v.fecha) as mes,
           SUM(CASE WHEN a.marca = {MARCA} THEN
               CASE WHEN v.operacion='+' THEN v.cantidad
                    WHEN v.operacion='-' THEN -v.cantidad ELSE 0 END
           ELSE 0 END) as topper_pares,
           SUM(CASE WHEN v.operacion='+' THEN v.cantidad
                    WHEN v.operacion='-' THEN -v.cantidad ELSE 0 END) as total_pares
    FROM msgestion01.dbo.omicron_ventas1 v
    LEFT JOIN msgestion01art.dbo.articulo a ON v.articulo = a.codigo
    WHERE v.codigo NOT IN {EXCL_CODIGOS}
      AND v.fecha >= '{anio}-01-01'
      AND v.fecha <= '{anio}-12-31'
      AND a.marca NOT IN {EXCL_MARCAS_GASTOS}
      AND a.rubro IN {RUBROS_CALZADO}
    GROUP BY MONTH(v.fecha)
    ORDER BY mes
    """
    cur = conn.cursor()
    cur.execute(sql)
    return {r.mes: (float(r.topper_pares or 0), float(r.total_pares or 0))
            for r in cur.fetchall()}


def query_stock_vel_total(conn):
    """Stock actual + vel_real total para calcular MOS."""
    sql = f"""
    SELECT
        (SELECT SUM(s.stock_actual)
         FROM msgestionC.dbo.stock s
         JOIN msgestion01art.dbo.articulo a ON s.articulo = a.codigo
         WHERE a.marca = {MARCA} AND s.deposito IN {DEPOSITOS}) as stock_total,
        (SELECT SUM(vr.vel_real)
         FROM omicronvt.dbo.vel_real_articulo vr
         JOIN (SELECT DISTINCT LEFT(codigo_sinonimo, 10) as csr
               FROM msgestion01art.dbo.articulo WHERE marca = {MARCA}) a2
           ON a2.csr = vr.codigo
         WHERE vr.vel_real > 0) as vel_real_total
    """
    cur = conn.cursor()
    cur.execute(sql)
    r = cur.fetchone()
    return float(r.stock_total or 0), float(r.vel_real_total or 1)


def query_top_verano(conn):
    """Top modelos Topper más vendidos en temporada PV (verano: Oct-Mar)."""
    sql = f"""
    SELECT TOP 15
           LEFT(a.codigo_sinonimo, 10) as csr,
           MIN(a.descripcion_1) as descripcion,
           sr.descripcion as subrubro_desc,
           SUM(CASE WHEN v1.operacion='+' THEN v1.cantidad
                    WHEN v1.operacion='-' THEN -v1.cantidad ELSE 0 END) as pares_pv,
           SUM(CASE WHEN MONTH(v2.fecha_comprobante) = 12 THEN
               CASE WHEN v1.operacion='+' THEN v1.cantidad
                    WHEN v1.operacion='-' THEN -v1.cantidad ELSE 0 END
               ELSE 0 END) as pares_dic
    FROM msgestionC.dbo.ventas1 v1
    JOIN msgestionC.dbo.ventas2 v2
        ON v1.codigo = v2.codigo AND v1.numero = v2.numero
        AND v1.letra = v2.letra AND v1.sucursal = v2.sucursal
    JOIN msgestion01art.dbo.articulo a ON v1.articulo = a.codigo
    LEFT JOIN msgestionC.dbo.subrubro sr ON a.subrubro = sr.codigo
    WHERE a.marca = {MARCA}
      AND v2.codigo NOT IN {EXCL_CODIGOS}
      AND (
        -- PV 2024/25: oct-dic 2024 + ene-mar 2025
        (v2.fecha_comprobante >= '2024-10-01' AND v2.fecha_comprobante < '2025-04-01')
        OR
        -- PV 2023/24: oct-dic 2023 + ene-mar 2024
        (v2.fecha_comprobante >= '2023-10-01' AND v2.fecha_comprobante < '2024-04-01')
      )
    GROUP BY LEFT(a.codigo_sinonimo, 10), sr.descripcion
    ORDER BY pares_pv DESC
    """
    cur = conn.cursor()
    cur.execute(sql)
    return cur.fetchall()


def query_so_mensual_historico_pv(conn):
    """SO mensual historico para meses PV (Oct-Mar) de los ultimos 3 años.
    Para mostrar tendencia de crecimiento verano.
    """
    sql = f"""
    SELECT YEAR(v.fecha) as anio,
           MONTH(v.fecha) as mes,
           SUM(CASE WHEN v.operacion='+' THEN v.cantidad
                    WHEN v.operacion='-' THEN -v.cantidad ELSE 0 END) as pares
    FROM msgestion01.dbo.omicron_ventas1 v
    LEFT JOIN msgestion01art.dbo.articulo a ON v.articulo = a.codigo
    WHERE a.marca = {MARCA}
      AND v.codigo NOT IN {EXCL_CODIGOS}
      AND v.fecha >= '2022-01-01'
      AND v.fecha <= '2025-12-31'
      AND MONTH(v.fecha) IN (1,2,3,10,11,12)
    GROUP BY YEAR(v.fecha), MONTH(v.fecha)
    ORDER BY anio, mes
    """
    cur = conn.cursor()
    cur.execute(sql)
    rows = {}
    for r in cur.fetchall():
        rows[(r.anio, r.mes)] = float(r.pares or 0)
    return rows


# ═════════════════════════════════════════════════════
# CONSTRUCCIÓN EXCEL
# ═════════════════════════════════════════════════════

def apply_header_style(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def cell_border(ws, row, col, value, fmt=None, bold=False, fill=None, align='center'):
    c = ws.cell(row=row, column=col, value=value)
    c.border = THIN_BORDER
    c.alignment = Alignment(horizontal=align)
    if fmt:
        c.number_format = fmt
    if bold:
        c.font = Font(bold=True)
    if fill:
        c.fill = fill
    return c


def mos_fill(mos):
    if mos < 2:
        return RED_FILL
    if mos < 4:
        return YELLOW_FILL
    return GREEN_FILL


def crear_hoja_so_share_mos(wb, so25, share25, stock_total, vel_real_total):
    """Hoja principal: SO/SHARE/MOS 2025 con highlight verano y colegial."""
    ws = wb.active
    ws.title = "Topper 2025"

    # ── Títulos ──
    ws.merge_cells('A1:N1')
    ws['A1'] = "HACHE CUATRO SRL — TOPPER — Sell Out 2025"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:N2')
    ws['A2'] = "Temporada Verano: Jul–Dic 2025 (entregas)  |  Pico Colegial: Febrero  |  Día del Niño: Agosto  |  Día de la Madre: Octubre"
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')

    # ── Leyenda colores ──
    ws.merge_cells('A3:N3')
    ws['A3'] = "  🔵 Temporada Verano (Jul-Dic)   🟡 Colegial (Feb)   🟠 Día del Niño (Ago)   🟣 Día de la Madre (Oct)   MOS: 🟢>4m  🟡2-4m  🔴<2m"
    ws['A3'].font = Font(italic=True, size=9, color="444444")

    # ── Header meses ──
    row = 5
    ws.column_dimensions['A'].width = 18
    cell_border(ws, row, 1, "MÉTRICA", bold=True, fill=HEADER_FILL, align='left')
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    for i, mes in enumerate(MESES_LABEL):
        col = i + 2
        c = cell_border(ws, row, col, mes, bold=True)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = 8
    cell_border(ws, row, 14, "TOTAL / PROM", bold=True)
    ws.cell(row=row, column=14).font = HEADER_FONT
    ws.cell(row=row, column=14).fill = HEADER_FILL
    ws.column_dimensions['N'].width = 12

    # ── Color de columnas verano en header ──
    for m in range(1, 13):
        col = m + 1
        if m in MESES_COLEGIAL:
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color="BF8F00", end_color="BF8F00", fill_type="solid")
        elif m in MESES_VERANO:
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color="1F5C99", end_color="1F5C99", fill_type="solid")

    # ── Fila SO pares ──
    row = 6
    cell_border(ws, row, 1, "SO (pares)", bold=True, align='left')
    total_pares = 0
    for m in range(1, 13):
        val = so25.get(m, 0)
        total_pares += val
        if m in MESES_COLEGIAL:
            fill = COLEGIAL_FILL
        elif m in MESES_DIA_NINO:
            fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")  # naranja
        elif m in MESES_DIA_MADRE:
            fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")  # violeta
        elif m in MESES_VERANO:
            fill = VERANO_FILL
        else:
            fill = None
        cell_border(ws, row, m + 1, int(val), fmt=NUM_FORMAT, fill=fill)
    cell_border(ws, row, 14, int(total_pares), fmt=NUM_FORMAT, bold=True, fill=TOTAL_FILL)

    # ── Fila SHARE pares ──
    row = 7
    cell_border(ws, row, 1, "SHARE (pares)", bold=True, align='left')
    shares_verano = []
    shares_total = []
    for m in range(1, 13):
        t, tot = share25.get(m, (0, 0))
        val = t / tot if tot > 0 else 0
        if m in MESES_COLEGIAL:
            fill = COLEGIAL_FILL
        elif m in MESES_DIA_NINO:
            fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
        elif m in MESES_DIA_MADRE:
            fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")
        elif m in MESES_VERANO:
            fill = VERANO_FILL
        else:
            fill = None
        cell_border(ws, row, m + 1, val, fmt=PCT_FORMAT, fill=fill)
        shares_total.append(val)
        if m in MESES_VERANO:
            shares_verano.append(val)
    prom_share = sum(shares_total) / len([s for s in shares_total if s > 0]) if shares_total else 0
    cell_border(ws, row, 14, prom_share, fmt=PCT_FORMAT, bold=True, fill=TOTAL_FILL)

    # ── Fila MOS (meses de stock) ──
    row = 8
    cell_border(ws, row, 1, "MOS (meses stock)", bold=True, align='left')
    vel_mensual = vel_real_total / 12 if vel_real_total > 0 else 1

    # MOS reconstruido mes a mes (aproximación: stock actual / vel promedio)
    # Para el proveedor mostramos MOS proyectado a futuro basado en stock actual
    # Los meses pasados se estiman por stock_inicio del mes (reconstrucción simple)
    mos_values = {}
    for m in range(1, 13):
        so_m = so25.get(m, 0)
        if vel_mensual > 0:
            mos = (stock_total / vel_mensual) * (1 + (m - 12) / 12) if m <= 12 else stock_total / vel_mensual
            # Simplificado: usar vel mensual promedio, ajustado por mes
            mos = stock_total / vel_mensual if m == 12 else so_m / vel_mensual * 2 if so_m > 0 else 0
        else:
            mos = 0
        mos_values[m] = mos

    # Mejor aproximación: usar SO del mes como proxy de demanda para calcular MOS ese mes
    for m in range(1, 13):
        so_m = so25.get(m, 0)
        # MOS = cuántos meses de stock teníamos ese mes = stock estimado / demanda mensual
        # Reconstruimos stock hacia atrás desde stock actual
        pass

    # Reconstrucción stock mes a mes hacia atrás desde stock actual
    stock_mes = {12: stock_total}
    for m in range(11, 0, -1):
        stock_mes[m] = stock_mes[m + 1] + so25.get(m + 1, 0)
    stock_mes[12] = stock_total  # stock actual al cierre 2025

    for m in range(1, 13):
        dem = so25.get(m, 0) if so25.get(m, 0) > 0 else vel_mensual
        mos = stock_mes.get(m, 0) / dem if dem > 0 else 0
        fill_mos = mos_fill(mos)
        if m in MESES_COLEGIAL:
            fill_mos = COLEGIAL_FILL if mos >= 2 else RED_FILL
        cell_border(ws, row, m + 1, round(mos, 1), fill=fill_mos)

    mos_actual = stock_total / vel_mensual if vel_mensual > 0 else 0
    cell_border(ws, row, 14, round(mos_actual, 1), bold=True, fill=mos_fill(mos_actual))

    # ── Totales verano destacados ──
    row = 10
    ws.merge_cells(f'A{row}:N{row}')
    ws[f'A{row}'] = "RESUMEN POR EVENTO COMERCIAL (Cono Sur)"
    ws[f'A{row}'].font = Font(bold=True, size=11, color="1F5C99")

    row = 11
    headers_resumen = ["Período / Evento", "SO Pares", "SHARE Pares", "% sobre año completo"]
    for i, h in enumerate(headers_resumen):
        cell_border(ws, row, i + 1, h, bold=True)
        ws.cell(row=row, column=i + 1).font = HEADER_FONT
        ws.cell(row=row, column=i + 1).fill = HEADER_FILL

    so_verano = sum(so25.get(m, 0) for m in MESES_VERANO)
    so_total_anio = sum(so25.get(m, 0) for m in range(1, 13))
    share_pares_verano = (sum(share25.get(m, (0, 0))[0] for m in MESES_VERANO) /
                          sum(share25.get(m, (0, 0))[1] for m in MESES_VERANO)
                          if sum(share25.get(m, (0, 0))[1] for m in MESES_VERANO) > 0 else 0)

    def share_mes(m):
        t, tot = share25.get(m, (0, 0))
        return t / tot if tot > 0 else 0

    row = 12
    datos = [
        ("Feb 2025 🎒 Colegial",        int(so25.get(2, 0)),  share_mes(2),  so25.get(2,0)/so_total_anio if so_total_anio else 0),
        ("Ago 2025 👦 Día del Niño",    int(so25.get(8, 0)),  share_mes(8),  so25.get(8,0)/so_total_anio if so_total_anio else 0),
        ("Oct 2025 👩 Día de la Madre", int(so25.get(10, 0)), share_mes(10), so25.get(10,0)/so_total_anio if so_total_anio else 0),
        ("Jul–Dic 2025 🔵 Verano",      int(so_verano),       share_pares_verano, so_verano/so_total_anio if so_total_anio else 0),
        ("TOTAL 2025",                  int(so_total_anio),   prom_share,    1.0),
    ]
    fills_resumen = [
        COLEGIAL_FILL,
        PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid"),
        VERANO_FILL,
        TOTAL_FILL,
    ]
    for i, (periodo, pares, share_val, pct) in enumerate(datos):
        fr = row + i
        cell_border(ws, fr, 1, periodo, bold=(i == 3), align='left')
        cell_border(ws, fr, 2, pares, fmt=NUM_FORMAT, bold=(i == 3))
        cell_border(ws, fr, 3, share_val, fmt=PCT_FORMAT, bold=(i == 3))
        cell_border(ws, fr, 4, pct, fmt=PCT_FORMAT, bold=(i == 3))
        if fills_resumen[i]:
            for col in range(1, 5):
                ws.cell(row=fr, column=col).fill = fills_resumen[i]

    # ── Ajuste columnas ──
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['N'].width = 14

    return ws


def crear_hoja_top_modelos(wb, top_rows):
    """Hoja 2: Top modelos verano — contexto interno para la reunión."""
    ws = wb.create_sheet("Top Modelos Verano")

    ws.merge_cells('A1:F1')
    ws['A1'] = "TOP MODELOS TOPPER — Temporada PV (Oct–Mar) — Últimas 2 temporadas"
    ws['A1'].font = TITLE_FONT

    ws.merge_cells('A2:F2')
    ws['A2'] = "Ordenados por ventas verano. Diciembre = pico colegial."
    ws['A2'].font = Font(italic=True, size=10, color="666666")

    row = 4
    headers = ["Modelo", "Categoría", "Pares PV (2 temp.)", "Pares Dic", "% en Dic"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    for i, r in enumerate(top_rows):
        fr = row + 1 + i
        pv = float(r.pares_pv or 0)
        dic = float(r.pares_dic or 0)
        pct_dic = dic / pv if pv > 0 else 0
        fill = COLEGIAL_FILL if pct_dic > 0.20 else None  # destaco si dic > 20% del PV
        cell_border(ws, fr, 1, r.descripcion or r.csr, align='left', fill=fill)
        cell_border(ws, fr, 2, r.subrubro_desc or "", align='left', fill=fill)
        cell_border(ws, fr, 3, int(pv), fmt=NUM_FORMAT, fill=fill)
        cell_border(ws, fr, 4, int(dic), fmt=NUM_FORMAT, fill=fill)
        cell_border(ws, fr, 5, pct_dic, fmt=PCT_FORMAT, fill=fill)

    ws.merge_cells(f'A{row+len(top_rows)+2}:F{row+len(top_rows)+2}')
    nota = ws.cell(row=row+len(top_rows)+2, column=1,
                   value="🟡 = modelo con >20% de ventas PV concentradas en diciembre (colegial)")
    nota.font = Font(italic=True, size=9, color="BF8F00")


def crear_hoja_historico_pv(wb, hist, so24, so25):
    """Hoja 3: Histórico PV (verano) por año — demuestra crecimiento."""
    ws = wb.create_sheet("Histórico Verano")

    ws.merge_cells('A1:G1')
    ws['A1'] = "HISTÓRICO VERANO (PV) — TOPPER H4 + CLZ"
    ws['A1'].font = TITLE_FONT

    ws.merge_cells('A2:G2')
    ws['A2'] = "Temporada PV = Oct-Mar. Incluye meses colegial (dic-ene)."
    ws['A2'].font = Font(italic=True, size=10, color="666666")

    # Construir tabla: temporadas PV 2022/23, 2023/24, 2024/25
    temporadas = [
        ("PV 2022/23", 2022, [10, 11, 12], 2023, [1, 2, 3]),
        ("PV 2023/24", 2023, [10, 11, 12], 2024, [1, 2, 3]),
        ("PV 2024/25", 2024, [10, 11, 12], 2025, [1, 2, 3]),
    ]

    row = 4
    headers = ["Temporada", "Oct", "Nov", "Dic ★", "Ene", "Feb", "Mar", "TOTAL PV"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 16
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 10

    for ti, (nombre, anio1, meses1, anio2, meses2) in enumerate(temporadas):
        fr = row + 1 + ti
        cell_border(ws, fr, 1, nombre, bold=True, align='left')
        total_pv = 0
        for ci, (a, m) in enumerate([(anio1, 10), (anio1, 11), (anio1, 12),
                                      (anio2, 1),  (anio2, 2),  (anio2, 3)]):
            val = int(hist.get((a, m), 0))
            total_pv += val
            fill = COLEGIAL_FILL if m == 12 else None
            cell_border(ws, fr, ci + 2, val, fmt=NUM_FORMAT, fill=fill)
        cell_border(ws, fr, 8, total_pv, fmt=NUM_FORMAT, bold=True, fill=TOTAL_FILL)

    # Crecimiento
    pv_totales = []
    for nombre, anio1, meses1, anio2, meses2 in temporadas:
        t = sum(hist.get((a, m), 0) for a, m in
                [(anio1, 10), (anio1, 11), (anio1, 12), (anio2, 1), (anio2, 2), (anio2, 3)])
        pv_totales.append(t)

    fr = row + len(temporadas) + 2
    ws.merge_cells(f'A{fr}:H{fr}')
    if len(pv_totales) >= 2 and pv_totales[0] > 0:
        crec = (pv_totales[-1] - pv_totales[0]) / pv_totales[0]
        ws[f'A{fr}'] = f"Crecimiento PV {temporadas[0][0]} → {temporadas[-1][0]}: {crec:+.0%}  |  Pares: {int(pv_totales[0]):,} → {int(pv_totales[-1]):,}"
        ws[f'A{fr}'].font = Font(bold=True, size=11, color="2F5496")


# ═════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════

def main():
    print("Conectando al ERP...")
    conn = get_conn()

    print("Leyendo SO 2025...")
    so25 = query_so_mensual_anio(conn, 2025)
    print("Leyendo SO 2024...")
    so24 = query_so_mensual_anio(conn, 2024)
    print("Leyendo SHARE 2025...")
    share25 = query_share_mensual_anio(conn, 2025)
    print("Leyendo stock + vel_real...")
    stock_total, vel_real_total = query_stock_vel_total(conn)
    print("Leyendo top modelos verano...")
    top_rows = query_top_verano(conn)
    print("Leyendo histórico PV...")
    hist = query_so_mensual_historico_pv(conn)

    conn.close()

    # Resumen consola
    so_anio = sum(so25.values())
    so_verano = sum(so25.get(m, 0) for m in [7,8,9,10,11,12])
    so_colegial = so25.get(12, 0)
    mos_actual = stock_total / (vel_real_total / 12) if vel_real_total > 0 else 0
    print(f"\n=== TOPPER 2025 ===")
    print(f"SO anual:         {int(so_anio):,} pares")
    print(f"SO verano Jul-Dic:{int(so_verano):,} pares ({so_verano/so_anio:.0%})")
    print(f"SO diciembre:     {int(so_colegial):,} pares")
    print(f"Stock actual:     {int(stock_total):,} pares")
    print(f"Vel real/mes:     {vel_real_total/12:.0f} pares")
    print(f"MOS actual:       {mos_actual:.1f} meses")

    print("\nGenerando Excel...")
    wb = Workbook()
    crear_hoja_so_share_mos(wb, so25, share25, stock_total, vel_real_total)
    crear_hoja_top_modelos(wb, top_rows)
    crear_hoja_historico_pv(wb, hist, so24, so25)

    out = os.path.join(os.path.dirname(__file__), "..", "_excel_pedidos",
                       "topper_preventa_verano2025_MANUEL.xlsx")
    out = os.path.normpath(out)
    wb.save(out)
    print(f"\n✓ Excel generado: {out}")
    print("  Hoja 1: Topper 2025     — SO/SHARE/MOS mensual con highlight verano")
    print("  Hoja 2: Top Modelos     — top 15 modelos verano (highlight colegial)")
    print("  Hoja 3: Histórico Verano— tendencia PV últimas 3 temporadas")


if __name__ == "__main__":
    main()
