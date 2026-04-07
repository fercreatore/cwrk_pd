#!/usr/bin/env python3
"""
insertar_atomik_inv26.py — Inserta 2 pedidos Atomik/VICBOR Invierno 2026
=========================================================================
PEDIDO 1: "ATOMIK INV 2026" (principal)
  Excel: /Volumes/compartido/COMPRAS/Pedidos Invierno/ATOMIK INV 2026.xlsx
  Sheets: MASCULINO, FEMENINO, KIDS (skip DISCIPLINA)
  74 SKUs, 1,078 pares, ~$45.5M
  Curva expansion: CURVA "40-43" + PARES POR TALLE "2,3,3,2" * PEDIDO multiplier

PEDIDO 2: "ATOMIK SUELTOS" (repo individual)
  Excel: /Volumes/compartido/COMPRAS/Pedidos Invierno/repo atomik sueltos.xlsx
  Sheet: SUELTOS
  28 lines, 68 pares, ~$1.8M
  Col V (22): descuento especial per-line (0-49.5%) → descuento_reng1

Empresa: H4 → INSERT en MSGESTION03.dbo.pedico2 + pedico1
Proveedor: 594 (VICBOR SRL / INDUSTRIAS AS S.A.)
Artículo matching: CODIGO ALFA → sinonimo en msgestion01art.dbo.articulo

EJECUTAR:
  python _scripts_oneshot/insertar_atomik_inv26.py --dry-run       (default)
  python _scripts_oneshot/insertar_atomik_inv26.py --ejecutar       (INSERT real)
  python _scripts_oneshot/insertar_atomik_inv26.py --solo-principal  (solo pedido 1)
  python _scripts_oneshot/insertar_atomik_inv26.py --solo-sueltos    (solo pedido 2)
"""

import sys
import os
import socket
from datetime import date, datetime

# ── Path setup ──
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl no está instalado. Ejecutar: pip install openpyxl")
    sys.exit(1)

import pyodbc

# ── Conexión ──
_hostname = socket.gethostname().upper()
if _hostname in ("DELL-SVR", "DELLSVR"):
    SERVIDOR = "localhost"
    DRIVER = "ODBC Driver 17 for SQL Server"
    EXTRAS = ""
else:
    SERVIDOR = "192.168.2.111"
    DRIVER = "ODBC Driver 17 for SQL Server"
    EXTRAS = "TrustServerCertificate=yes;Encrypt=no;"

def get_conn(base):
    return (
        f"DRIVER={{{DRIVER}}};"
        f"SERVER={SERVIDOR};"
        f"DATABASE={base};"
        f"UID=am;PWD=dl;"
        f"{EXTRAS}"
    )

# ── Constantes ──
EMPRESA      = "H4"
BASE_INSERT  = "MSGESTION03"   # H4 → msgestion03
PROVEEDOR    = 594
DENOMINACION = "INDUSTRIAS AS S.A."
DESCUENTO_PROV = 20   # descuento proveedor VICBOR (config.py)
DESCUENTO_BON  = 0

EXCEL_PRINCIPAL = "/Volumes/compartido/COMPRAS/Pedidos Invierno/ATOMIK INV 2026.xlsx"
EXCEL_SUELTOS   = "/Volumes/compartido/COMPRAS/Pedidos Invierno/repo atomik sueltos.xlsx"

SHEETS_PRINCIPAL = ["MASCULINO", "FEMENINO", "KIDS"]


# =============================================================================
# UTILIDADES
# =============================================================================

def buscar_articulo_por_sinonimo(cursor, sinonimo):
    """
    Busca un artículo por sinonimo en msgestion01art.dbo.articulo.
    Retorna (codigo, descripcion) o (None, None).
    """
    sinonimo = str(sinonimo).strip()
    if not sinonimo:
        return None, None

    # Búsqueda exacta primero
    cursor.execute(
        "SELECT codigo, descripcion_1 FROM msgestion01art.dbo.articulo "
        "WHERE codigo_sinonimo = ? AND estado = 'V'",
        (sinonimo,)
    )
    row = cursor.fetchone()
    if row:
        return row[0], (row[1] or "").strip()

    # Si no matchea exacto, probar con LIKE (por si tiene espacios)
    cursor.execute(
        "SELECT codigo, descripcion_1 FROM msgestion01art.dbo.articulo "
        "WHERE codigo_sinonimo LIKE ? AND estado = 'V'",
        (sinonimo + '%',)
    )
    row = cursor.fetchone()
    if row:
        return row[0], (row[1] or "").strip()

    return None, None


def parsear_curva(curva_str):
    """
    Parsea una curva de talles como "40-43" → [40, 41, 42, 43]
    o "35-40" → [35, 36, 37, 38, 39, 40]
    """
    curva_str = str(curva_str).strip()
    if '-' in curva_str:
        partes = curva_str.split('-')
        inicio = int(partes[0].strip())
        fin = int(partes[1].strip())
        return list(range(inicio, fin + 1))
    # Si es un solo número
    return [int(curva_str)]


def parsear_pares_por_talle(ppt_str):
    """
    Parsea distribución "2,3,3,2" → [2, 3, 3, 2]
    """
    ppt_str = str(ppt_str).strip()
    return [int(x.strip()) for x in ppt_str.split(',') if x.strip()]


def safe_float(val, default=0.0):
    """Convierte a float de forma segura."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    """Convierte a int de forma segura."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


# =============================================================================
# PEDIDO 1: ATOMIK INV 2026 (principal con curvas)
# =============================================================================

def leer_pedido_principal(filepath):
    """
    Lee el Excel principal de Atomik INV 2026.
    Sheets: MASCULINO, FEMENINO, KIDS
    Headers row 2, data from row 3.
    Col H (8): CODIGO ALFA (sinonimo)
    Col I (9): PEDIDO (multiplier 0/1/2)
    Col E (5): CURVA (rango talles "40-43")
    Col F (6): PARES POR TALLE (distribución "2,3,3,2")
    Col K (11): PRECIO MAYORISTA
    Col T (20): PARES total por módulo

    Returns list of dicts with expanded talle lines.
    """
    if not os.path.exists(filepath):
        print(f"ERROR: No se encuentra el archivo: {filepath}")
        return []

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    items = []
    skus_vistos = 0

    for sheet_name in SHEETS_PRINCIPAL:
        if sheet_name not in wb.sheetnames:
            print(f"  WARN: Sheet '{sheet_name}' no encontrada en el Excel, skip")
            continue

        ws = wb[sheet_name]
        print(f"  Leyendo sheet: {sheet_name}")
        filas_procesadas = 0

        for row in ws.iter_rows(min_row=3, values_only=True):
            # row es tuple, indices 0-based
            # Col H (8) = index 7, Col I (9) = index 8
            # Col E (5) = index 4, Col F (6) = index 5
            # Col K (11) = index 10, Col T (20) = index 19
            if len(row) < 11:
                continue

            codigo_alfa = row[7]   # Col H - CODIGO ALFA (sinonimo)
            pedido_mult = safe_int(row[8], 0)  # Col I - PEDIDO multiplier
            curva_str   = row[4]   # Col E - CURVA
            ppt_str     = row[5]   # Col F - PARES POR TALLE
            precio_may  = safe_float(row[10], 0)  # Col K - PRECIO MAYORISTA
            pares_total = safe_int(row[19] if len(row) > 19 else 0)  # Col T - PARES

            if not codigo_alfa or pedido_mult <= 0:
                continue

            codigo_alfa = str(codigo_alfa).strip()
            if not curva_str or not ppt_str:
                print(f"    WARN: {codigo_alfa} tiene PEDIDO={pedido_mult} pero sin CURVA/PPT, skip")
                continue

            skus_vistos += 1
            talles = parsear_curva(curva_str)
            distrib = parsear_pares_por_talle(ppt_str)

            if len(talles) != len(distrib):
                print(f"    WARN: {codigo_alfa} — curva {curva_str} tiene {len(talles)} talles "
                      f"pero PPT '{ppt_str}' tiene {len(distrib)} valores. Skip.")
                continue

            # Expandir: un item por talle
            for talle, qty_base in zip(talles, distrib):
                qty = qty_base * pedido_mult
                if qty <= 0:
                    continue
                # Construir sinonimo con talle: codigo_alfa + talle en formato que matchee
                # El sinonimo del artículo en el ERP ya incluye el talle
                # Para Atomik: sinonimo base + sufijo talle (ej: "25211313441IA" + "40")
                sino_con_talle = f"{codigo_alfa}{talle}"
                items.append({
                    "sinonimo": sino_con_talle,
                    "sinonimo_base": codigo_alfa,
                    "talle": str(talle),
                    "cantidad": qty,
                    "precio": precio_may,
                    "descuento_reng1": DESCUENTO_PROV,
                    "descuento_reng2": DESCUENTO_BON,
                    "sheet": sheet_name,
                    "pedido_mult": pedido_mult,
                })

            filas_procesadas += 1

        print(f"    {filas_procesadas} SKUs con PEDIDO>0 en {sheet_name}")

    wb.close()
    print(f"  Total SKUs con pedido: {skus_vistos}")
    print(f"  Total líneas expandidas (por talle): {len(items)}")
    return items


# =============================================================================
# PEDIDO 2: ATOMIK SUELTOS (repo individual)
# =============================================================================

def leer_pedido_sueltos(filepath):
    """
    Lee el Excel de sueltos/repo Atomik.
    Sheet: SUELTOS (primera hoja o por nombre)
    Col L (12) = index 11: CODIGO ALFA (sinonimo con talle incluido)
    Col F (6) = index 5: TALLE
    Col M (13) = index 12: PEDIDO qty (directo, no multiplier)
    Col P (16) = index 15: PRECIO MAYORISTA
    Col V (22) = index 21: DESCUENTO_ESPECIAL (0-49.5%)

    Returns list of dicts.
    """
    if not os.path.exists(filepath):
        print(f"ERROR: No se encuentra el archivo: {filepath}")
        return []

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Buscar sheet SUELTOS o usar la primera
    if "SUELTOS" in wb.sheetnames:
        ws = wb["SUELTOS"]
    else:
        ws = wb.active
        print(f"  WARN: Sheet 'SUELTOS' no encontrada, usando sheet activa: {ws.title}")

    items = []
    print(f"  Leyendo sheet: {ws.title}")

    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 16:
            continue

        codigo_alfa = row[11]    # Col L - CODIGO ALFA (con talle)
        talle       = row[5]     # Col F - TALLE
        pedido_qty  = safe_int(row[12], 0)  # Col M - PEDIDO
        precio_may  = safe_float(row[15], 0)  # Col P - PRECIO MAYORISTA
        desc_esp    = safe_float(row[21] if len(row) > 21 else 0, 0)  # Col V - DESCUENTO ESPECIAL

        if not codigo_alfa or pedido_qty <= 0:
            continue

        codigo_alfa = str(codigo_alfa).strip()

        # desc_esp puede venir como porcentaje (49.5) o como decimal (0.495)
        if desc_esp > 1:
            desc_pct = desc_esp  # ya es porcentaje
        else:
            desc_pct = desc_esp * 100  # convertir de decimal a porcentaje

        items.append({
            "sinonimo": codigo_alfa,
            "sinonimo_base": codigo_alfa,
            "talle": str(talle).strip() if talle else "",
            "cantidad": pedido_qty,
            "precio": precio_may,
            "descuento_reng1": round(desc_pct, 2),
            "descuento_reng2": DESCUENTO_BON,
            "sheet": "SUELTOS",
            "pedido_mult": 1,
        })

    wb.close()
    print(f"  Total líneas con PEDIDO>0: {len(items)}")
    return items


# =============================================================================
# RESOLVER ARTÍCULOS (sinonimo → codigo ERP)
# =============================================================================

def resolver_articulos(cursor, items):
    """
    Para cada item, busca el artículo por sinonimo.
    Retorna (matched, unmatched).
    """
    matched = []
    unmatched = []

    # Cache para evitar queries repetidas
    cache = {}

    for item in items:
        sino = item["sinonimo"]
        if sino in cache:
            cod, desc = cache[sino]
        else:
            cod, desc = buscar_articulo_por_sinonimo(cursor, sino)
            cache[sino] = (cod, desc)

        if cod:
            item["articulo"] = cod
            item["descripcion"] = desc
            item["codigo_sinonimo"] = sino
            matched.append(item)
        else:
            unmatched.append(item)

    return matched, unmatched


# =============================================================================
# INSERTAR PEDIDO
# =============================================================================

def insertar_pedido(conn, items, observaciones, dry_run=True):
    """
    Inserta un pedido en pedico2 (cabecera) + pedico1 (detalle).
    items: lista de dicts con articulo, descripcion, cantidad, precio, etc.
    """
    cursor = conn.cursor()

    # ── Próximo número y orden ──
    cursor.execute(
        f"SELECT MAX(CAST(numero AS INT)) FROM {BASE_INSERT}.dbo.pedico2 "
        f"WHERE codigo=8 AND letra='X' AND sucursal=1"
    )
    row = cursor.fetchone()
    ultimo_num = row[0] if row[0] else 0
    nuevo_numero = ultimo_num + 1

    cursor.execute(
        f"SELECT MAX(orden) FROM {BASE_INSERT}.dbo.pedico2 "
        f"WHERE codigo=8 AND letra='X' AND sucursal=1"
    )
    row = cursor.fetchone()
    ultimo_orden = row[0] if row[0] else 0
    nuevo_orden = (ultimo_orden + 1) if (ultimo_orden + 1) <= 99 else 1

    hoy = date.today()
    ahora = datetime.now()
    total_pares = sum(i["cantidad"] for i in items)
    total_monto = sum(i["cantidad"] * i["precio"] for i in items)

    # ── Buscar nombre proveedor en BD ──
    cursor.execute(
        f"SELECT denominacion FROM {BASE_INSERT}.dbo.proveedores WHERE numero=?",
        (PROVEEDOR,)
    )
    row = cursor.fetchone()
    nombre_prov = (row[0] or DENOMINACION).strip() if row else DENOMINACION

    # ── Imprimir resumen ──
    print(f"\n{'='*70}")
    print(f"PEDIDO A INSERTAR:")
    print(f"  Número:      {nuevo_numero}")
    print(f"  Orden:       {nuevo_orden}")
    print(f"  Proveedor:   {PROVEEDOR} — {nombre_prov}")
    print(f"  Empresa:     {EMPRESA} → {BASE_INSERT}")
    print(f"  Fecha:       {hoy}")
    print(f"  Pares:       {total_pares}")
    print(f"  Monto bruto: ${total_monto:,.2f}")
    print(f"  Líneas:      {len(items)}")
    print(f"  Obs:         {observaciones}")
    print(f"{'='*70}")

    for i, item in enumerate(items, 1):
        desc_r1 = item.get("descuento_reng1", 0)
        desc_str = f" desc={desc_r1}%" if desc_r1 > 0 else ""
        print(f"  {i:>3}. [{item['articulo']}] {item['descripcion'][:50]:50s} "
              f"x{item['cantidad']:>3}  ${item['precio']:>12,.2f}{desc_str}")

    print(f"\n  TOTAL: {total_pares} pares | ${total_monto:,.2f}")

    if dry_run:
        print(f"\n[DRY RUN] No se ejecutó nada.")
        return nuevo_numero

    # ── INSERT pedico2 (cabecera) ──
    sql_cab = f"""
    INSERT INTO {BASE_INSERT}.dbo.pedico2
        (codigo, letra, sucursal, numero, orden,
         cuenta, denominacion, fecha_comprobante, fecha_proceso,
         observaciones, estado, usuario, importe_neto,
         descuento_general, monto_descuento,
         bonificacion_general, monto_bonificacion,
         financiacion_general, monto_financiacion,
         iva1, monto_iva1, iva2, monto_iva2, monto_impuesto,
         monto_exento, deposito,
         zona, condicion_iva, numero_cuit, copias,
         cuenta_y_orden, pack, reintegro, cambio, transferencia,
         entregador, campo, sistema_cc, moneda, sector,
         forma_pago, plan_canje, tipo_vcto_pago, tipo_operacion, tipo_ajuste,
         medio_pago, cuenta_cc, concurso)
    VALUES
        (8, 'X', 1, ?, ?,
         ?, ?, ?, ?,
         ?, 'V', 'COWORK', ?,
         0, 0, 0, 0, 0, 0,
         21, 0, 10.5, 0, 0,
         0, 0,
         4, 'I', '', 1,
         'N', 'N', 'N', 'N', 'N',
         0, 0, 2, 0, 0,
         0, 'N', 0, 0, 0,
         ' ', ?, 'N')
    """
    cursor.execute(sql_cab, (
        nuevo_numero, nuevo_orden,
        PROVEEDOR, nombre_prov, hoy, ahora,
        observaciones, total_monto,
        PROVEEDOR,  # cuenta_cc
    ))
    print(f"\npedico2 insertado: numero={nuevo_numero}, orden={nuevo_orden}")

    # ── INSERT pedico1 (detalle) ──
    sql_det = f"""
    INSERT INTO {BASE_INSERT}.dbo.pedico1
        (codigo, letra, sucursal, numero, orden, renglon,
         articulo, descripcion, codigo_sinonimo,
         cantidad, precio,
         descuento_reng1, descuento_reng2,
         cuenta, fecha, fecha_entrega,
         estado, usuario)
    VALUES
        (8, 'X', 1, ?, ?, ?,
         ?, ?, ?,
         ?, ?,
         ?, ?,
         ?, ?, ?,
         'V', 'COWORK')
    """
    for i, item in enumerate(items, 1):
        cursor.execute(sql_det, (
            nuevo_numero, nuevo_orden, i,
            item["articulo"],
            item["descripcion"],
            item.get("codigo_sinonimo", ""),
            item["cantidad"],
            item["precio"],
            item.get("descuento_reng1", DESCUENTO_PROV),
            item.get("descuento_reng2", DESCUENTO_BON),
            PROVEEDOR,
            hoy,
            hoy,  # fecha_entrega = hoy (ajustar si hay fecha de entrega pactada)
        ))

    conn.commit()
    print(f"\nPedido insertado exitosamente: #{nuevo_numero}")
    print(f"  {len(items)} líneas | {total_pares} pares | ${total_monto:,.2f}")
    return nuevo_numero


# =============================================================================
# MAIN
# =============================================================================

def main():
    args = sys.argv[1:]
    dry_run = "--ejecutar" not in args
    solo_principal = "--solo-principal" in args
    solo_sueltos = "--solo-sueltos" in args

    if dry_run:
        print("\n*** MODO DRY RUN — para ejecutar: python insertar_atomik_inv26.py --ejecutar ***\n")
    else:
        print("\n*** MODO EJECUCION REAL — se escribira en la base ***")
        resp = input("    Confirmar? (s/N): ").strip().lower()
        if resp != "s":
            print("    Cancelado.")
            sys.exit(0)

    # ── Conexión ──
    conn_art = pyodbc.connect(get_conn("msgestion01art"))
    cursor_art = conn_art.cursor()

    conn_insert = pyodbc.connect(get_conn(BASE_INSERT.lower()))
    conn_insert.autocommit = False

    # ====================================================================
    # PEDIDO 1: ATOMIK INV 2026 (principal)
    # ====================================================================
    if not solo_sueltos:
        print("\n" + "=" * 70)
        print("PEDIDO 1: ATOMIK INV 2026 (principal)")
        print("=" * 70)

        items_principal = leer_pedido_principal(EXCEL_PRINCIPAL)

        if items_principal:
            matched, unmatched = resolver_articulos(cursor_art, items_principal)

            if unmatched:
                print(f"\n*** {len(unmatched)} ARTICULOS SIN MATCH (necesitan alta) ***")
                sinos_vistos = set()
                for u in unmatched:
                    if u["sinonimo"] not in sinos_vistos:
                        sinos_vistos.add(u["sinonimo"])
                        print(f"  FALTA: sinonimo={u['sinonimo']}  talle={u['talle']}  "
                              f"qty={u['cantidad']}  precio={u['precio']:.2f}  "
                              f"sheet={u['sheet']}")
                print(f"  Total sinónimos sin match: {len(sinos_vistos)}")
                pares_perdidos = sum(u["cantidad"] for u in unmatched)
                print(f"  Pares que no se insertaran: {pares_perdidos}")

            if matched:
                total_p = sum(m["cantidad"] for m in matched)
                total_m = sum(m["cantidad"] * m["precio"] for m in matched)
                print(f"\n  Artículos matcheados: {len(matched)} líneas, {total_p} pares, ${total_m:,.2f}")

                obs1 = f"Pedido Atomik INV 2026. {total_p} pares. ${total_m:,.0f}. VICBOR SRL"
                try:
                    num1 = insertar_pedido(conn_insert, matched, obs1, dry_run=dry_run)
                    print(f"\n  Pedido principal: #{num1}")
                except Exception as e:
                    conn_insert.rollback()
                    print(f"\n  ERROR insertando pedido principal: {e}")
                    if not dry_run:
                        raise
            else:
                print("\n  NINGÚN artículo matcheó. Revisar sinónimos en el Excel vs ERP.")
        else:
            print("  No se encontraron items en el Excel principal.")

    # ====================================================================
    # PEDIDO 2: ATOMIK SUELTOS (repo)
    # ====================================================================
    if not solo_principal:
        print("\n" + "=" * 70)
        print("PEDIDO 2: ATOMIK SUELTOS / REPO")
        print("=" * 70)

        items_sueltos = leer_pedido_sueltos(EXCEL_SUELTOS)

        if items_sueltos:
            matched2, unmatched2 = resolver_articulos(cursor_art, items_sueltos)

            if unmatched2:
                print(f"\n*** {len(unmatched2)} ARTICULOS SIN MATCH (necesitan alta) ***")
                for u in unmatched2:
                    print(f"  FALTA: sinonimo={u['sinonimo']}  talle={u['talle']}  "
                          f"qty={u['cantidad']}  precio={u['precio']:.2f}")
                pares_perdidos2 = sum(u["cantidad"] for u in unmatched2)
                print(f"  Pares que no se insertaran: {pares_perdidos2}")

            if matched2:
                total_p2 = sum(m["cantidad"] for m in matched2)
                total_m2 = sum(m["cantidad"] * m["precio"] for m in matched2)
                print(f"\n  Artículos matcheados: {len(matched2)} líneas, {total_p2} pares, ${total_m2:,.2f}")

                obs2 = f"Pedido Atomik Sueltos/Repo. {total_p2} pares. ${total_m2:,.0f}. VICBOR SRL"
                try:
                    num2 = insertar_pedido(conn_insert, matched2, obs2, dry_run=dry_run)
                    print(f"\n  Pedido sueltos: #{num2}")
                except Exception as e:
                    conn_insert.rollback()
                    print(f"\n  ERROR insertando pedido sueltos: {e}")
                    if not dry_run:
                        raise
            else:
                print("\n  NINGÚN artículo matcheó. Revisar sinónimos en el Excel vs ERP.")
        else:
            print("  No se encontraron items en el Excel de sueltos.")

    # ── Cleanup ──
    conn_art.close()
    conn_insert.close()

    print("\n" + "=" * 70)
    print("FIN")
    print("=" * 70)
    if dry_run:
        print("Para insertar de verdad: python _scripts_oneshot/insertar_atomik_inv26.py --ejecutar")


if __name__ == "__main__":
    main()
