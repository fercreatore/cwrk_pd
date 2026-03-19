"""
Script para enviar WhatsApp masivo a agencias de viaje con el banner de Valijas GO.

Funcionalidades:
  - Lee contactos de LISTA_AGENCIAS_WHATSAPP.xlsx
  - Genera links wa.me con mensaje personalizado
  - Convierte FLYER_AGENCIAS_V2.html a imagen PNG para adjuntar
  - Abre WhatsApp Web automáticamente (uno por uno, con delay para no ser baneado)
  - Registra envíos en el Excel (Fecha_1er_msj, Estado)

Uso:
    python3 valijas/enviar_whatsapp.py              # Ver lista de contactos
    python3 valijas/enviar_whatsapp.py --enviar      # Abrir wa.me para cada contacto
    python3 valijas/enviar_whatsapp.py --links       # Solo generar links (sin abrir)
    python3 valijas/enviar_whatsapp.py --tipo General  # Filtrar por tipo
    python3 valijas/enviar_whatsapp.py --ciudad Rosario  # Filtrar por ciudad
"""

import openpyxl
import re
import os
import sys
import webbrowser
import time
import subprocess
from datetime import datetime
from urllib.parse import quote

# Rutas
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "LISTA_AGENCIAS_WHATSAPP.xlsx")
FLYER_HTML = os.path.join(SCRIPT_DIR, "FLYER_AGENCIAS_V2.html")
FLYER_PNG = os.path.join(SCRIPT_DIR, "FLYER_AGENCIAS_V2.png")
IMAGENES_DIR = os.path.join(SCRIPT_DIR, "imagenes")

# Fotos de sets para mandar después del primer mensaje
FOTOS_SETS = {
    "negro": "7981DFA3-BA8D-4DF0-827C-93262386D1A4_1_105_c.jpeg",
    "rosa_gold": "BE724906-E30F-4FA6-8B5D-75637C010B7E_1_105_c.jpeg",
    "rosa": "442EA850-5101-45F6-BE50-11F7644F0ED6_1_105_c.jpeg",
    "rojo": "31A6CE80-E550-46CF-B026-725D2C610AE5.jpeg",
    "interior": "195644CA-8CD3-4EDE-8E7B-25B92455EEDF.jpeg",
}

# Número propio para recibir pedidos
WHATSAPP_PROPIO = "5493462676300"

# Rate limiting — WhatsApp puede banear si mandás más de 15-20/hora
DELAY_ENTRE_MENSAJES = 45  # segundos entre cada apertura de wa.me
MAX_POR_SESION = 15  # máximo por sesión para no ser baneado


def normalizar_telefono(telefono):
    """Normalizar número a formato internacional sin +."""
    if not telefono:
        return None
    # Limpiar
    limpio = re.sub(r'[^\d]', '', str(telefono))
    # Si empieza con 0, agregar 54
    if limpio.startswith("0"):
        limpio = "54" + limpio[1:]
    # Si es un 0800/0810, no es WhatsApp
    if limpio.startswith("08"):
        return None
    # Si tiene menos de 10 dígitos, no es válido
    if len(limpio) < 10:
        return None
    return limpio


def get_mensaje(nombre_agencia, tipo="General"):
    """Generar mensaje personalizado según tipo de agencia."""

    if tipo == "Estudiantil":
        mensaje = f"""Hola! Buenas tardes

Soy Fernando de Calzalindo. Te escribo porque tenemos algo ideal para tus grupos de egresados.

Trajimos un contenedor de valijas rigidas importadas marca GO. Sets de 3 valijas (17" + 19" + 21"), perfectas para el viaje de egresados.

La 17" entra en low cost (Flybondi/JetSmart)
La 19" para cabina estandar (Aerolineas/LATAM)
La 21" para despachar

La propuesta es simple:
- Vos lo compartis en el grupo de WhatsApp del viaje
- Por cada set vendido te damos $20.000 de comision
- Precio exclusivo grupo: $179.999 (en la web esta $199.999)
- Envio gratis a todo el pais
- 6 cuotas sin interes

Un grupo de 40 egresados donde 15 compren = $300.000 de comision para vos.
Y si manejas 5 grupos... ya haces las cuentas.

Nosotros nos encargamos de TODO: atencion, envio, postventa.

Te mando las fotos y el mensaje armado para que copies y pegues en el grupo?"""

    elif tipo == "Grupal":
        mensaje = f"""Hola! Buenas tardes

Soy Fernando de Calzalindo, tenemos 20 anios en el rubro calzado y accesorios en Venado Tuerto.

Te escribo porque vi que {nombre_agencia} organiza viajes grupales y tenemos una propuesta que les puede servir.

Importamos directo de fabrica un contenedor de valijas rigidas marca GO. Son sets de 3 (17" + 19" + 21"), 8 ruedas 360, ABS, cierre combinacion. Vienen en 4 colores.

La propuesta:
- Vos compartis un mensaje en tus grupos de viaje
- Por cada set que se venda, te damos $20.000 de comision
- El precio para tus pasajeros: $179.999 (10% menos que la web)
- Nosotros nos encargamos de envio, stock, atencion, todo
- Envio gratis a todo el pais

Si un grupo de 30 pasajeros y 10 compran = $200.000 para vos sin hacer nada.

Te mando fotos?"""

    else:  # General
        mensaje = f"""Hola! Buenas tardes

Soy Fernando de Calzalindo, tenemos 20 anios en el rubro calzado y accesorios en Venado Tuerto.

Te escribo porque tenemos una propuesta para los pasajeros de {nombre_agencia} que puede ser negocio para los dos.

Importamos directo de fabrica un contenedor de valijas rigidas marca GO. Son sets de 3 (17" + 19" + 21"), 8 ruedas 360, ABS, cierre combinacion. Vienen en 4 colores.

La propuesta:
- Vos compartis un mensaje en tus grupos de viaje
- Por cada set que se venda, te damos $20.000 de comision
- El precio para tus pasajeros: $179.999 (10% menos que la web)
- Nosotros nos encargamos de envio, stock, atencion, todo
- Envio gratis a todo el pais

Si un grupo de 30 pasajeros y 10 compran = $200.000 para vos sin hacer nada.

Te mando fotos?"""

    return mensaje


def generar_link_wame(telefono, mensaje):
    """Generar link wa.me con mensaje pre-cargado."""
    return f"https://wa.me/{telefono}?text={quote(mensaje)}"


def leer_contactos(filtro_tipo=None, filtro_ciudad=None, solo_pendientes=True):
    """Leer contactos del Excel y filtrar."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    contactos = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = [cell.value for cell in row]
        nombre = vals[0]
        if not nombre:
            continue

        contacto = {
            "nombre": vals[0],
            "ciudad": vals[1],
            "tipo": vals[2] or "General",
            "telefono": vals[3],
            "email": vals[4],
            "web": vals[5],
            "estado": vals[6] or "Pendiente",
            "fecha_1er_msj": vals[7],
            "respondio": vals[8],
            "sets_vendidos": vals[9],
            "comision_pagada": vals[10],
            "row_num": row[0].row,  # para actualizar después
        }

        # Filtros
        if solo_pendientes and contacto["estado"] not in ("Pendiente", None):
            continue
        if filtro_tipo and contacto["tipo"] != filtro_tipo:
            continue
        if filtro_ciudad and filtro_ciudad.lower() not in (contacto["ciudad"] or "").lower():
            continue

        contactos.append(contacto)

    wb.close()
    return contactos


def marcar_enviado(row_num, estado="Enviado"):
    """Actualizar estado y fecha en el Excel."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Columna G = Estado (7), Columna H = Fecha_1er_msj (8)
    ws.cell(row=row_num, column=7, value=estado)
    ws.cell(row=row_num, column=8, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    wb.save(EXCEL_PATH)
    wb.close()


def convertir_flyer_a_png():
    """Convertir el flyer HTML a PNG para enviar por WhatsApp."""
    if os.path.exists(FLYER_PNG):
        print(f"  Flyer PNG ya existe: {FLYER_PNG}")
        return FLYER_PNG

    # Intentar con wkhtmltoimage (si está instalado)
    try:
        subprocess.run(
            ["wkhtmltoimage", "--width", "1080", FLYER_HTML, FLYER_PNG],
            capture_output=True, timeout=30
        )
        if os.path.exists(FLYER_PNG):
            print(f"  Flyer convertido a PNG: {FLYER_PNG}")
            return FLYER_PNG
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass

    # Intentar con screenshot via Selenium (fallback)
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        opts = Options()
        opts.add_argument("--headless")
        opts.add_argument("--window-size=1080,1920")
        driver = webdriver.Chrome(options=opts)
        driver.get(f"file://{FLYER_HTML}")
        driver.save_screenshot(FLYER_PNG)
        driver.quit()
        print(f"  Flyer screenshot: {FLYER_PNG}")
        return FLYER_PNG
    except Exception:
        pass

    print("  NOTA: No se pudo convertir el flyer a PNG automáticamente.")
    print("  Opciones:")
    print("    1. Abrir FLYER_AGENCIAS_V2.html en Chrome y hacer screenshot")
    print("    2. Instalar wkhtmltoimage: brew install wkhtmltopdf")
    print("    3. Usar FLYER_VALIJAS_GO.png que ya existe")

    # Fallback al PNG existente
    fallback = os.path.join(SCRIPT_DIR, "FLYER_VALIJAS_GO.png")
    if os.path.exists(fallback):
        print(f"  Usando fallback: FLYER_VALIJAS_GO.png")
        return fallback

    return None


def copiar_imagenes_al_portapapeles(foto_key):
    """Copiar imagen al portapapeles en macOS."""
    foto = FOTOS_SETS.get(foto_key)
    if not foto:
        return False
    path = os.path.join(IMAGENES_DIR, foto)
    if not os.path.exists(path):
        return False
    try:
        subprocess.run(
            ["osascript", "-e", f'set the clipboard to (read (POSIX file "{path}") as JPEG picture)'],
            capture_output=True, timeout=5
        )
        return True
    except Exception:
        return False


def mostrar_contactos(contactos):
    """Mostrar tabla de contactos."""
    con_wa = [c for c in contactos if normalizar_telefono(c["telefono"])]
    sin_wa = [c for c in contactos if not normalizar_telefono(c["telefono"])]

    print(f"\n{'='*80}")
    print(f"LISTA DE AGENCIAS — {len(contactos)} total, {len(con_wa)} con WhatsApp")
    print(f"{'='*80}")

    print(f"\n{'#':>3}  {'Nombre':<30} {'Ciudad':<15} {'Tipo':<12} {'WhatsApp':<16} {'Estado'}")
    print("-" * 95)

    for i, c in enumerate(con_wa, 1):
        tel = normalizar_telefono(c["telefono"])
        print(f"{i:>3}  {c['nombre']:<30} {c['ciudad']:<15} {c['tipo']:<12} {tel:<16} {c['estado']}")

    if sin_wa:
        print(f"\n--- Sin WhatsApp ({len(sin_wa)}) ---")
        for c in sin_wa:
            print(f"     {c['nombre']:<30} {c['ciudad']:<15} {c['tipo']:<12} {'SIN NUMERO':<16} (web: {c.get('web', '-')})")


def ejecutar_envio(contactos, dry_run=False):
    """Ejecutar envío de mensajes vía wa.me."""
    con_wa = [c for c in contactos if normalizar_telefono(c["telefono"])]

    if not con_wa:
        print("\nNo hay contactos con WhatsApp para enviar.")
        return

    print(f"\n{'='*80}")
    print(f"ENVÍO WHATSAPP — {len(con_wa)} contactos")
    print(f"Delay: {DELAY_ENTRE_MENSAJES}s entre mensajes | Max por sesión: {MAX_POR_SESION}")
    print(f"{'='*80}")

    # Convertir flyer
    flyer_path = convertir_flyer_a_png()

    enviados = 0
    for i, c in enumerate(con_wa):
        if enviados >= MAX_POR_SESION:
            print(f"\n  Límite de {MAX_POR_SESION} mensajes alcanzado. Esperá 1 hora antes de continuar.")
            break

        tel = normalizar_telefono(c["telefono"])
        mensaje = get_mensaje(c["nombre"], c["tipo"])
        link = generar_link_wame(tel, mensaje)

        print(f"\n[{i+1}/{len(con_wa)}] {c['nombre']} ({c['ciudad']}, {c['tipo']})")
        print(f"  Tel: {tel}")
        print(f"  Link: {link[:80]}...")

        if dry_run:
            print("  [DRY RUN — no se abre]")
        else:
            # Abrir wa.me en el navegador
            webbrowser.open(link)
            print("  Abierto en WhatsApp Web")

            # Marcar como enviado
            marcar_enviado(c["row_num"], "Enviado")
            print("  Excel actualizado: Enviado")

            enviados += 1

            if i < len(con_wa) - 1 and enviados < MAX_POR_SESION:
                print(f"  Esperando {DELAY_ENTRE_MENSAJES}s...")
                time.sleep(DELAY_ENTRE_MENSAJES)

    print(f"\n{'='*80}")
    print(f"RESUMEN: {enviados} mensajes {'generados' if dry_run else 'enviados'}")
    if flyer_path:
        print(f"Banner: {flyer_path}")
        print("NOTA: El banner hay que adjuntarlo manualmente en cada chat de WhatsApp Web")
    print(f"{'='*80}")


def generar_links(contactos):
    """Solo generar e imprimir links wa.me."""
    con_wa = [c for c in contactos if normalizar_telefono(c["telefono"])]

    print(f"\n{'='*80}")
    print(f"LINKS WA.ME — {len(con_wa)} contactos con WhatsApp")
    print(f"{'='*80}\n")

    for c in con_wa:
        tel = normalizar_telefono(c["telefono"])
        mensaje = get_mensaje(c["nombre"], c["tipo"])
        link = generar_link_wame(tel, mensaje)
        print(f"{c['nombre']} ({c['tipo']}):")
        print(f"  {link}")
        print()


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Envío WhatsApp masivo a agencias — Valijas GO")
    parser.add_argument("--enviar", action="store_true", help="Abrir wa.me para cada contacto")
    parser.add_argument("--links", action="store_true", help="Solo generar links wa.me")
    parser.add_argument("--tipo", type=str, help="Filtrar por tipo (General, Grupal, Estudiantil)")
    parser.add_argument("--ciudad", type=str, help="Filtrar por ciudad")
    parser.add_argument("--todos", action="store_true", help="Incluir contactos ya enviados")
    parser.add_argument("--flyer", action="store_true", help="Solo convertir flyer a PNG")
    args = parser.parse_args()

    if args.flyer:
        convertir_flyer_a_png()
        return

    contactos = leer_contactos(
        filtro_tipo=args.tipo,
        filtro_ciudad=args.ciudad,
        solo_pendientes=not args.todos,
    )

    if args.links:
        generar_links(contactos)
    elif args.enviar:
        ejecutar_envio(contactos)
    else:
        mostrar_contactos(contactos)
        print(f"\nPara enviar: python3 valijas/enviar_whatsapp.py --enviar")
        print(f"Solo links:  python3 valijas/enviar_whatsapp.py --links")
        print(f"Filtrar:     python3 valijas/enviar_whatsapp.py --tipo Estudiantil --ciudad Rosario")


if __name__ == "__main__":
    main()
