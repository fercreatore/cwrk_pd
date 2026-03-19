"""
API local que sirve contactos + mensajes personalizados como JSON para n8n.

n8n consume este endpoint via HTTP Request node y luego envía por WhatsApp Business Cloud.

Uso:
    python3 valijas/api_contactos_whatsapp.py

Endpoints:
    GET http://localhost:5050/contactos              → todos los contactos pendientes con mensaje
    GET http://localhost:5050/contactos?tipo=Estudiantil  → filtrar por tipo
    GET http://localhost:5050/contactos?ciudad=Rosario    → filtrar por ciudad
    POST http://localhost:5050/marcar_enviado         → {row_num: 5, estado: "Enviado"}
    GET http://localhost:5050/flyer                   → URL del flyer HTML
"""

import json
import re
import os
import openpyxl
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "LISTA_AGENCIAS_WHATSAPP.xlsx")

WHATSAPP_PROPIO = "5493462676300"
PORT = 5050


def normalizar_telefono(telefono):
    if not telefono:
        return None
    limpio = re.sub(r'[^\d]', '', str(telefono))
    if limpio.startswith("0"):
        limpio = "54" + limpio[1:]
    if limpio.startswith("08"):
        return None
    if len(limpio) < 10:
        return None
    return limpio


def get_mensaje(nombre_agencia, tipo="General"):
    if tipo == "Estudiantil":
        return (
            f"Hola! Buenas tardes\n\n"
            f"Soy Fernando de Calzalindo. Te escribo porque tenemos algo ideal para tus grupos de egresados.\n\n"
            f"Trajimos un contenedor de valijas rigidas importadas marca GO. Sets de 3 valijas (17\" + 19\" + 21\"), "
            f"perfectas para el viaje de egresados.\n\n"
            f"La 17\" entra en low cost (Flybondi/JetSmart)\n"
            f"La 19\" para cabina estandar (Aerolineas/LATAM)\n"
            f"La 21\" para despachar\n\n"
            f"La propuesta es simple:\n"
            f"- Vos lo compartis en el grupo de WhatsApp del viaje\n"
            f"- Por cada set vendido te damos $20.000 de comision\n"
            f"- Precio exclusivo grupo: $179.999 (en la web esta $199.999)\n"
            f"- Envio gratis a todo el pais\n"
            f"- 6 cuotas sin interes\n\n"
            f"Un grupo de 40 egresados donde 15 compren = $300.000 de comision para vos.\n\n"
            f"Nosotros nos encargamos de TODO: atencion, envio, postventa.\n\n"
            f"Te mando las fotos y el mensaje armado para que copies y pegues en el grupo?"
        )
    elif tipo == "Grupal":
        return (
            f"Hola! Buenas tardes\n\n"
            f"Soy Fernando de Calzalindo, tenemos 20 anios en el rubro calzado y accesorios en Venado Tuerto.\n\n"
            f"Te escribo porque vi que {nombre_agencia} organiza viajes grupales y tenemos una propuesta "
            f"que les puede servir.\n\n"
            f"Importamos directo de fabrica un contenedor de valijas rigidas marca GO. "
            f"Son sets de 3 (17\" + 19\" + 21\"), 8 ruedas 360, ABS, cierre combinacion. Vienen en 4 colores.\n\n"
            f"La propuesta:\n"
            f"- Vos compartis un mensaje en tus grupos de viaje\n"
            f"- Por cada set que se venda, te damos $20.000 de comision\n"
            f"- El precio para tus pasajeros: $179.999 (10% menos que la web)\n"
            f"- Nosotros nos encargamos de envio, stock, atencion, todo\n"
            f"- Envio gratis a todo el pais\n\n"
            f"Si un grupo de 30 pasajeros y 10 compran = $200.000 para vos sin hacer nada.\n\n"
            f"Te mando fotos?"
        )
    else:
        return (
            f"Hola! Buenas tardes\n\n"
            f"Soy Fernando de Calzalindo, tenemos 20 anios en el rubro calzado y accesorios en Venado Tuerto.\n\n"
            f"Te escribo porque tenemos una propuesta para los pasajeros de {nombre_agencia} "
            f"que puede ser negocio para los dos.\n\n"
            f"Importamos directo de fabrica un contenedor de valijas rigidas marca GO. "
            f"Son sets de 3 (17\" + 19\" + 21\"), 8 ruedas 360, ABS, cierre combinacion. Vienen en 4 colores.\n\n"
            f"La propuesta:\n"
            f"- Vos compartis un mensaje en tus grupos de viaje\n"
            f"- Por cada set que se venda, te damos $20.000 de comision\n"
            f"- El precio para tus pasajeros: $179.999 (10% menos que la web)\n"
            f"- Nosotros nos encargamos de envio, stock, atencion, todo\n"
            f"- Envio gratis a todo el pais\n\n"
            f"Si un grupo de 30 pasajeros y 10 compran = $200.000 para vos sin hacer nada.\n\n"
            f"Te mando fotos?"
        )


def leer_contactos(filtro_tipo=None, filtro_ciudad=None, solo_pendientes=True):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    contactos = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = [cell.value for cell in row]
        nombre = vals[0]
        if not nombre:
            continue

        telefono = normalizar_telefono(vals[3])
        if not telefono:
            continue

        estado = vals[6] or "Pendiente"
        tipo = vals[2] or "General"
        ciudad = vals[1] or ""

        if solo_pendientes and estado not in ("Pendiente", None):
            continue
        if filtro_tipo and tipo != filtro_tipo:
            continue
        if filtro_ciudad and filtro_ciudad.lower() not in ciudad.lower():
            continue

        mensaje = get_mensaje(nombre, tipo)

        contactos.append({
            "row_num": row[0].row,
            "nombre": nombre,
            "ciudad": ciudad,
            "tipo": tipo,
            "telefono": telefono,
            "email": vals[4],
            "web": vals[5],
            "estado": estado,
            "mensaje": mensaje,
            "wa_me_link": f"https://wa.me/{telefono}",
        })

    wb.close()
    return contactos


def marcar_enviado(row_num, estado="Enviado"):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.cell(row=row_num, column=7, value=estado)
    ws.cell(row=row_num, column=8, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    wb.save(EXCEL_PATH)
    wb.close()


class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)

        if parsed.path == "/contactos":
            params = parse_qs(parsed.query)
            tipo = params.get("tipo", [None])[0]
            ciudad = params.get("ciudad", [None])[0]
            todos = params.get("todos", ["false"])[0] == "true"

            contactos = leer_contactos(
                filtro_tipo=tipo,
                filtro_ciudad=ciudad,
                solo_pendientes=not todos,
            )

            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps(contactos, ensure_ascii=False, indent=2).encode())

        elif parsed.path == "/flyer":
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            flyer_url = f"file://{os.path.join(SCRIPT_DIR, 'FLYER_AGENCIAS_V2.html')}"
            self.wfile.write(json.dumps({"flyer_html": flyer_url}).encode())

        elif parsed.path == "/health":
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"status": "ok", "timestamp": datetime.now().isoformat()}).encode())

        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        if self.path == "/marcar_enviado":
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length)) if length else {}

            row_num = body.get("row_num")
            estado = body.get("estado", "Enviado")

            if row_num:
                marcar_enviado(int(row_num), estado)
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.end_headers()
                self.wfile.write(json.dumps({"ok": True, "row": row_num}).encode())
            else:
                self.send_response(400)
                self.end_headers()
                self.wfile.write(b'{"error": "row_num requerido"}')
        else:
            self.send_response(404)
            self.end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def log_message(self, format, *args):
        print(f"  [{datetime.now().strftime('%H:%M:%S')}] {args[0]}")


def main():
    print(f"{'='*60}")
    print(f"API Contactos WhatsApp — Valijas GO")
    print(f"{'='*60}")
    print(f"Servidor: http://localhost:{PORT}")
    print(f"Excel: {EXCEL_PATH}")
    print()
    print(f"Endpoints para n8n:")
    print(f"  GET  http://localhost:{PORT}/contactos")
    print(f"  GET  http://localhost:{PORT}/contactos?tipo=Estudiantil")
    print(f"  GET  http://localhost:{PORT}/contactos?ciudad=Rosario")
    print(f"  POST http://localhost:{PORT}/marcar_enviado  (body: {{row_num: 5}})")
    print(f"  GET  http://localhost:{PORT}/health")
    print()

    contactos = leer_contactos()
    print(f"Contactos pendientes con WhatsApp: {len(contactos)}")
    for c in contactos:
        print(f"  {c['nombre']:<30} {c['ciudad']:<15} {c['tipo']:<12} {c['telefono']}")
    print()
    print("Esperando requests de n8n...")
    print()

    server = HTTPServer(("0.0.0.0", PORT), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServidor detenido.")
        server.server_close()


if __name__ == "__main__":
    main()
